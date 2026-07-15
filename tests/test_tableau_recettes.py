# -*- coding: utf-8 -*-
"""Tests du tableau Excel « recettes à définir avec les chefs » :
classification des produits (fabriqués / composés / revente / inactifs),
génération du classeur, puis import du fichier rempli (aller-retour complet)."""
import json

import pandas as pd
import pytest
from openpyxl import load_workbook

import outils.generer_tableau_recettes as G
import outils.importer_recettes_chefs as I
from paul_forecast import config


FIN = pd.Timestamp("2026-06-28")


@pytest.fixture
def ventes(monkeypatch):
    """Ventes synthétiques couvrant chaque statut attendu."""
    monkeypatch.setitem(config.BOM, "PRODUIT EXACT TEST", {"Farine de blé T65 (g)": 100.0})
    rows = []
    recents = pd.date_range(FIN - pd.Timedelta(days=89), FIN)
    for d in recents:
        rows.append({"Date": d, "Code": 1, "Produit": "FLUTE TEST 250GR",
                     "Famille": "BOULANGERIE", "Quantite": 10.0, "CA_TTC": 0.0})
        rows.append({"Date": d, "Code": 2, "Produit": "PRODUIT MYSTERE CUISINE",
                     "Famille": "CUISINE", "Quantite": 4.0, "CA_TTC": 0.0})
        rows.append({"Date": d, "Code": 3, "Produit": "MENU TEST 1",
                     "Famille": "MENU", "Quantite": 6.0, "CA_TTC": 0.0})
        rows.append({"Date": d, "Code": 4, "Produit": "COCA COLA 33CL",
                     "Famille": "BEVERAGE", "Quantite": 20.0, "CA_TTC": 0.0})
        rows.append({"Date": d, "Code": 5, "Produit": "PRODUIT EXACT TEST",
                     "Famille": "BOULANGERIE", "Quantite": 8.0, "CA_TTC": 0.0})
        rows.append({"Date": d, "Code": 6, "Produit": "COFFRET TEST GOURMAND",
                     "Famille": "CUISINE", "Quantite": 3.0, "CA_TTC": 0.0})
        rows.append({"Date": d, "Code": 8, "Produit": "PDJ TEST EXPRESS",
                     "Famille": "Autres", "Quantite": 5.0, "CA_TTC": 0.0})
        rows.append({"Date": d, "Code": 9, "Produit": "MINI CROISS PDJ ENFANT TEST",
                     "Famille": "Autres", "Quantite": 5.0, "CA_TTC": 0.0})
        # mono-ingrédient / stock (pas de recette)
        rows.append({"Date": d, "Code": 10, "Produit": "SUPPLEMENT FROMAGE TEST",
                     "Famille": "CUISINE", "Quantite": 5.0, "CA_TTC": 0.0})
        rows.append({"Date": d, "Code": 11, "Produit": "POT GLACE 1 BOULE TEST",
                     "Famille": "CUISINE", "Quantite": 5.0, "CA_TTC": 0.0})
        rows.append({"Date": d, "Code": 12, "Produit": "ACCOMPAGNEMENT RIZ TEST",
                     "Famille": "CUISINE", "Quantite": 5.0, "CA_TTC": 0.0})
    # produit inactif : dernière vente il y a > 90 jours
    rows.append({"Date": FIN - pd.Timedelta(days=200), "Code": 7, "Produit": "VIEUX PRODUIT",
                 "Famille": "BOULANGERIE", "Quantite": 50.0, "CA_TTC": 0.0})
    return pd.DataFrame(rows)


def _classif(ventes):
    return G.classifier(G.stats_produits(ventes)).set_index("Produit")["Statut"]


def test_classification(ventes):
    s = _classif(ventes)
    assert s["FLUTE TEST 250GR"] == "a_definir"          # fabriqué, recette estimée
    assert s["PRODUIT MYSTERE CUISINE"] == "a_definir"   # fabriqué, aucune recette
    assert s["PRODUIT EXACT TEST"] == "exacte"           # déjà défini
    assert s["MENU TEST 1"] == "exclu_compose"           # famille MENU
    assert s["COFFRET TEST GOURMAND"] == "exclu_compose" # mot COFFRET
    assert s["PDJ TEST EXPRESS"] == "exclu_compose"      # formule petit-déjeuner (préfixe)
    assert s["MINI CROISS PDJ ENFANT TEST"] == "a_definir"  # vrai produit (PDJ pas en préfixe)
    assert s["COCA COLA 33CL"] == "exclu_revente"        # marque revendue
    assert s["VIEUX PRODUIT"] == "exclu_inactif"         # < 5 ventes en 90 j
    assert s["SUPPLEMENT FROMAGE TEST"] == "mono_stock"  # supplément
    assert s["POT GLACE 1 BOULE TEST"] == "mono_stock"   # boule de glace
    assert s["ACCOMPAGNEMENT RIZ TEST"] == "mono_stock"  # accompagnement


def test_generation_classeur(ventes, tmp_path):
    chemin = str(tmp_path / "recettes.xlsx")
    G.generer(chemin, df_ventes=ventes)
    wb = load_workbook(chemin)
    # une feuille par catégorie (dans l'ordre : familles connues puis inconnues),
    # encadrée par les feuilles fixes ; les composés (MENU) ont leur feuille aussi
    assert wb.sheetnames == ["Mode d'emploi", "BOULANGERIE", "CUISINE", "Autres",
                             "MENU", "Mono-ingrédient (stock)", "Revendus (sans recette)"]
    mono = [r[0] for r in wb["Mono-ingrédient (stock)"].iter_rows(min_row=2, values_only=True)]
    assert {"SUPPLEMENT FROMAGE TEST", "POT GLACE 1 BOULE TEST",
            "ACCOMPAGNEMENT RIZ TEST"} <= set(mono)

    par_prod = {}
    for nom_ws in ("BOULANGERIE", "CUISINE", "Autres"):
        cur = None
        for r in wb[nom_ws].iter_rows(min_row=2, values_only=True):
            if r[0]:
                cur = r[0]                    # le nom n'est que sur la ligne-titre
            if cur:
                par_prod.setdefault(cur, []).append(r)
    # flûte (feuille BOULANGERIE) : estimation pré-remplie (farine…) + lignes vides
    fl = par_prod["FLUTE TEST 250GR"]
    assert any("Farine" in str(r[3]) for r in fl)
    assert any(r[3] in (None, "") for r in fl)                    # lignes à compléter
    # chaque catégorie ne contient QUE ses produits
    assert "FLUTE TEST 250GR" in [r[0] for r in wb["BOULANGERIE"].iter_rows(min_row=2, values_only=True)]
    assert "PRODUIT MYSTERE CUISINE" in [r[0] for r in wb["CUISINE"].iter_rows(min_row=2, values_only=True)]
    # produit sans recette : uniquement des lignes vides (6)
    my = par_prod["PRODUIT MYSTERE CUISINE"]
    assert len(my) == G.LIGNES_VIDES_SANS_ESTIMATION
    assert all(r[3] in (None, "") for r in my)
    # exclusions : ni coca ni mono-ingrédient dans les feuilles de saisie ;
    # le menu est dans SA feuille (MENU), pas dans les autres catégories
    assert "MENU TEST 1" not in par_prod and "COCA COLA 33CL" not in par_prod
    assert "POT GLACE 1 BOULE TEST" not in par_prod and "SUPPLEMENT FROMAGE TEST" not in par_prod
    menu = [r[0] for r in wb["MENU"].iter_rows(min_row=2, values_only=True)]
    assert "MENU TEST 1" in menu
    revendus = [r[0] for r in wb["Revendus (sans recette)"].iter_rows(min_row=2, values_only=True)]
    assert "COCA COLA 33CL" in revendus
    # le composé (COFFRET, famille CUISINE) reste corrigeable dans sa catégorie
    assert "COFFRET TEST GOURMAND" in par_prod
    # produit à recette exacte : listé dans sa catégorie, recette pré-remplie
    # depuis recettes_exactes.json avec une origine explicite
    pe = par_prod["PRODUIT EXACT TEST"]
    assert any("Farine de blé T65" in str(r[3]) for r in pe)
    assert any("exacte" in str(r[6]).lower() or "recette" in str(r[6]).lower() for r in pe)


def test_import_apres_remplissage(ventes, tmp_path):
    chemin = str(tmp_path / "recettes.xlsx")
    G.generer(chemin, df_ventes=ventes)

    # ── simulation du remplissage par le chef (feuille BOULANGERIE) ──────────
    wb = load_workbook(chemin)
    ws = wb["BOULANGERIE"]
    ligne_vide, en_flute = None, False
    for row in ws.iter_rows(min_row=2):
        if row[0].value == "FLUTE TEST 250GR":
            en_flute = True
            row[2].value = "OUI"              # validation sur la ligne-titre (col C)
        if not en_flute:
            continue
        if str(row[3].value or "").startswith("Farine"):
            row[4].value = 200                # corrige la quantité estimée
        elif str(row[3].value or "").startswith("Levure"):
            row[4].value = 0                  # 0 = ingrédient supprimé
        elif not row[3].value and ligne_vide is None:
            ligne_vide = row
    ligne_vide[3].value = "Levain"            # nouvel ingrédient, en kg
    ligne_vide[4].value = 0.05
    ligne_vide[5].value = "kg"
    autre = [r for r in ws.iter_rows(min_row=2)
             if not r[3].value and r is not ligne_vide
             and r[0].row > ligne_vide[0].row][0]
    autre[3].value = "Truc bizarre"
    autre[4].value = 3
    autre[5].value = "tasse"                  # unité invalide -> anomalie
    wb.save(chemin)

    # ── import ───────────────────────────────────────────────────────────────
    rj, rp = str(tmp_path / "recettes_exactes.json"), str(tmp_path / "prov.json")
    rapport = I.importer(chemin, chemin_json=rj, chemin_prov=rp)
    assert rapport["importes"] == 1 and rapport["produits"] == ["FLUTE TEST 250GR"]
    assert any("tasse" in a for a in rapport["anomalies"])

    recettes = json.load(open(rj, encoding="utf-8"))
    fl = recettes["FLUTE TEST 250GR"]
    assert fl["Farine de blé T55 (g)"] == 200.0
    assert fl["Levain (g)"] == 50.0                       # 0.05 kg -> 50 g
    assert not any(k.startswith("Levure") for k in fl)    # quantité 0 = retiré
    assert "Eau (ml)" in fl                               # estimation conservée telle quelle
    assert "Truc bizarre" not in " ".join(fl)             # unité invalide ignorée
    # le produit non validé n'est pas importé
    assert "PRODUIT MYSTERE CUISINE" not in recettes
    prov = json.load(open(rp, encoding="utf-8"))
    assert prov["FLUTE TEST 250GR"]["source"] == "tableau chefs (xlsx)"


def test_import_sans_validation(ventes, tmp_path):
    chemin = str(tmp_path / "recettes.xlsx")
    G.generer(chemin, df_ventes=ventes)
    rj = str(tmp_path / "recettes_exactes.json")
    rapport = I.importer(chemin, chemin_json=rj, chemin_prov=str(tmp_path / "p.json"))
    assert rapport["importes"] == 0
    import os
    assert not os.path.exists(rj)             # rien écrit si rien de validé


def test_estimateur_recettes_riches():
    import outils.estim_recettes as E
    # crêpe banane chocolat : base pâte + garniture spécifique
    c = E.estimer("CREPE BANANE CHOCO", "CUISINE")
    assert "Farine de blé T45 (g)" in c and "Œufs (g)" in c
    assert "Banane (g)" in c and "Pâte à tartiner chocolat (g)" in c
    # salade niçoise : base + thon/œufs/olives
    s = E.estimer("SALADE NICOISE", "CUISINE")
    assert "Thon (g)" in s and "Olives (g)" in s and "Salade verte (g)" in s
    # cappuccino : café + lait + gobelet
    caf = E.estimer("CAPPUCINO VAE", "BEVERAGE")
    assert caf.get("Café en grains (g)") == 9 and "Lait entier (ml)" in caf
    # quiche lorraine
    q = E.estimer("Mini quiche lorraine", "CUISINE")
    assert "Pâte brisée (g)" in q and "Lardons (g)" in q
    # le pain nu reste géré ailleurs -> estimateur ne renvoie rien
    assert E.estimer("FLUTE 250GR", "BOULANGERIE") == {}
    # une recette-type a plusieurs ingrédients (plus riche qu'avant)
    h = E.estimer("HAMBOURGEOIS GOURMAND", "CUISINE")
    assert len(h) >= 4 and "Pain hamburger (g)" in h
