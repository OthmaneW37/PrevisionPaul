# -*- coding: utf-8 -*-
"""
Restitution des résultats : tableaux de bord (gérant + matières premières),
bon de commande textuel et export Excel multi-feuilles.
"""

import os
import re

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt

from . import config
from .logging_setup import get_logger

logger = get_logger()


# ==============================================================================
# UTILITAIRES INGRÉDIENTS
# ==============================================================================
def obtenir_categorie_ingredient(ing):
    boulangerie  = ["farine", "levure", "sel", "eau", "graines"]
    vienn_patiss = ["beurre", "feuilletée", "choux", "chocolat", "raisins", "crème",
                    "fraises", "tarte", "éclair"]
    traiteur     = ["baguette", "jambon", "thon", "poulet", "parmesan", "mayonnaise",
                    "sauce", "salade"]
    boissons     = ["café", "oranges", "sucre"]

    ing_lower = str(ing).lower()
    if any(m in ing_lower for m in boulangerie):
        return "Boulangerie"
    if any(m in ing_lower for m in vienn_patiss):
        return "Viennoiserie & Pâtisserie"
    if any(m in ing_lower for m in traiteur):
        return "Traiteur & Sandwicherie"
    if any(m in ing_lower for m in boissons):
        return "Boissons & Bar"
    return "Emballages & Consommables"


def formater_quantite(ing, val):
    if "(g)" in ing and val >= 1000:
        return f"{val / 1000:.2f} kg"
    if "(ml)" in ing and val >= 1000:
        return f"{val / 1000:.2f} L"
    if any(u in ing.lower() for u in ["unité", "tranche", "sachet", "paille",
                                      "feuille", "boîte", "gobelet"]):
        return f"{np.ceil(val):.0f} pcs"
    return f"{val:.2f}"


# ==============================================================================
# BON DE COMMANDE TEXTUEL
# ==============================================================================
def generer_rapport_approvisionnement(df_besoins_mrp, output_dir):
    """Génère un bon de commande texte clair pour la prochaine période de planification."""
    if len(df_besoins_mrp) == 0:
        logger.info("Aucun besoin en ingrédient calculé.")
        return

    prochaine_date = df_besoins_mrp['Date'].min()
    date_str = str(prochaine_date)[:10]

    df_prochain = df_besoins_mrp[df_besoins_mrp['Date'] == prochaine_date].copy()
    df_prochain['Categorie'] = df_prochain['Ingredient'].apply(obtenir_categorie_ingredient)

    lignes = []
    lignes.append("=" * 70)
    lignes.append(" BON DE COMMANDE ET PLAN D'APPROVISIONNEMENT - BOULANGERIE PAUL")
    lignes.append(f" Periode cible : {date_str}")
    lignes.append("=" * 70)
    lignes.append("\nEn tant que gerant de la boutique PAUL, voici les volumes d'achat")
    lignes.append("a prevoir pour couvrir vos besoins de production sur la periode :\n")

    categories = ["Boulangerie", "Viennoiserie & Patisserie", "Traiteur & Sandwicherie",
                  "Boissons & Bar", "Emballages & Consommables"]
    cat_map = {"Viennoiserie & P\u00e2tisserie": "Viennoiserie & Patisserie"}
    df_prochain['Categorie'] = df_prochain['Categorie'].replace(cat_map)

    for cat in categories:
        df_cat = df_prochain[df_prochain['Categorie'] == cat]
        if len(df_cat) == 0:
            continue
        lignes.append(f"[{cat.upper()}]")
        lignes.append("-" * 45)
        for _, row in df_cat.sort_values(by='Quantite_Requise', ascending=False).iterrows():
            ing = row['Ingredient']
            val_format = formater_quantite(ing, row['Quantite_Requise'])
            lignes.append(f"  - {ing:<35} : {val_format}")
        lignes.append("")

    lignes.append("=" * 70)
    rapport_path = os.path.join(output_dir, "bon_de_commande_gerant.txt")
    lignes.append(f" Rapport sauvegarde sous : {rapport_path}")
    lignes.append("=" * 70)

    texte_rapport = "\n".join(lignes)
    os.makedirs(output_dir, exist_ok=True)
    with open(rapport_path, "w", encoding="utf-8") as f:
        f.write(texte_rapport)
    logger.info("Bon de commande sauvegardé dans : '%s'", rapport_path)


# ==============================================================================
# DASHBOARD GÉRANT (4 panneaux)
# ==============================================================================
def tracer_dashboard_mrp(df_hist_total, df_prev_total, df_ingredients_fc,
                         dict_prevision_prod, produits_uniques, output_dir):
    """Tableau de bord gérant : CA global, CA par produit, top volumes et top CA."""
    fig = plt.figure(figsize=(18, 13))
    fig.patch.set_facecolor('#f8f9fa')
    plt.style.use('seaborn-v0_8-whitegrid' if 'seaborn-v0_8-whitegrid' in plt.style.available else 'default')

    couleurs_produits = [
        '#2ecc71', '#3498db', '#e74c3c', '#f39c12', '#9b59b6',
        '#1abc9c', '#e67e22', '#c0392b', '#2980b9', '#27ae60',
        '#d35400', '#8e44ad', '#16a085'
    ]

    # --- PANNEAU 1 : CA MENSUEL GLOBAL ---
    ax1 = fig.add_subplot(2, 2, 1)
    ax1.fill_between(df_hist_total['Date'], df_hist_total['Chiffre_Affaires_Total'],
                     alpha=0.15, color='#2c3e50')
    ax1.plot(df_hist_total['Date'], df_hist_total['Chiffre_Affaires_Total'],
             label='Historique Reel', color='#2c3e50', linewidth=2.5, marker='o', markersize=4)

    derniere_date_hist = df_hist_total['Date'].iloc[-1]
    derniere_val_hist  = df_hist_total['Chiffre_Affaires_Total'].iloc[-1]
    dates_fc = pd.concat([pd.Series([derniere_date_hist]), df_prev_total['Date']]).reset_index(drop=True)
    col_principale = 'Rev_Prev_Holt_Winters' if 'Rev_Prev_Holt_Winters' in df_prev_total.columns else 'Rev_Prev_Moyenne_Mobile'
    vals_fc = pd.concat([pd.Series([derniere_val_hist]), df_prev_total[col_principale]]).reset_index(drop=True)

    MAPE_HW = 0.143
    vals_prev_np = df_prev_total[col_principale].values.astype(float)
    n_fc = len(vals_prev_np)
    spread_coef = np.linspace(1.0, 1.5, n_fc)
    ci_lower = np.concatenate([[derniere_val_hist], vals_prev_np * (1 - MAPE_HW * spread_coef)])
    ci_upper = np.concatenate([[derniere_val_hist], vals_prev_np * (1 + MAPE_HW * spread_coef)])
    ax1.fill_between(dates_fc, ci_lower, ci_upper, alpha=0.18, color='#e67e22',
                     label=f'IC ±{MAPE_HW*100:.0f}% (MAPE backtest)')

    ax1.fill_between(dates_fc, vals_fc, alpha=0.10, color='#e67e22')
    ax1.plot(dates_fc, vals_fc, label='Prévision principale (Holt-Winters)', color='#e67e22',
             linewidth=2.5, linestyle='--', marker='s', markersize=4)

    vals_decompo = pd.concat([pd.Series([derniere_val_hist]),
                              df_prev_total['Rev_Prev_Decompo_Saisonniere']]).reset_index(drop=True)
    ax1.plot(dates_fc, vals_decompo, label='Décompo Saisonnière (référence)', color='#8e44ad',
             linewidth=1.5, linestyle=':', marker=None, alpha=0.6)

    if config.ACTIVER_PROPHET and 'Rev_Prev_Prophet' in df_prev_total.columns:
        vals_prophet = pd.concat([pd.Series([derniere_val_hist]),
                                  df_prev_total['Rev_Prev_Prophet']]).reset_index(drop=True)
        ax1.plot(dates_fc, vals_prophet, label='Prevision (Prophet)', color='#e67e22',
                 linewidth=2.0, linestyle=':', marker='^', markersize=4)

    ymax_ax1 = max(float(df_hist_total['Chiffre_Affaires_Total'].max()),
                   float(df_prev_total['Rev_Prev_Decompo_Saisonniere'].max()))
    dates_fc_dt = pd.to_datetime(dates_fc)
    for fete in config.FETES_MAROCAINES:
        debut_f = pd.Timestamp(fete["debut"])
        fin_f   = pd.Timestamp(fete["fin"])
        if debut_f <= dates_fc_dt.iloc[-1] and fin_f >= dates_fc_dt.iloc[0]:
            mid_f = debut_f + (fin_f - debut_f) / 2
            ax1.axvspan(debut_f, fin_f, alpha=0.08, color='#e74c3c', zorder=0)
            ax1.text(mid_f, ymax_ax1 * 0.97, fete["nom"][:12],
                     fontsize=6.5, color='#c0392b', ha='center', va='top', rotation=90)

    ax1.set_title("Chiffre d'Affaires mensuel (tous produits)", fontsize=11, fontweight='bold', pad=8)
    ax1.set_ylabel("CA (MAD)", fontsize=9)
    ax1.legend(fontsize=8, loc='upper left')
    ax1.grid(True, linestyle=':', alpha=0.5)
    ax1.tick_params(axis='x', rotation=30, labelsize=8)
    ax1.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f'{x:,.0f} MAD'))

    # --- PANNEAU 2 : CA PRÉVISIONNEL PAR PRODUIT (Top 6) ---
    ax2 = fig.add_subplot(2, 2, 2)
    ca_par_produit = {}
    for prod in produits_uniques:
        vals_rev = dict_prevision_prod[prod]['Rev_Prev_Decompo_Saisonniere'].values
        ca_par_produit[prod] = float(np.nansum(vals_rev))
    top6 = sorted(ca_par_produit, key=ca_par_produit.get, reverse=True)[:6]

    n_periodes = len(df_prev_total)
    x_idx = np.arange(n_periodes)
    labels_dates = [pd.to_datetime(d).strftime('%b %Y') for d in df_prev_total['Date'].values]

    bottom = np.zeros(n_periodes)
    for idx_p, prod in enumerate(top6):
        vals_rev = dict_prevision_prod[prod]['Rev_Prev_Decompo_Saisonniere'].values.astype(float)
        vals_rev = vals_rev[:n_periodes] if len(vals_rev) >= n_periodes else np.pad(vals_rev, (0, n_periodes - len(vals_rev)))
        vals_rev = np.nan_to_num(vals_rev, nan=0.0)
        ax2.bar(x_idx, vals_rev, bottom=bottom, label=prod[:22],
                color=couleurs_produits[idx_p], width=0.7, alpha=0.85)
        bottom += vals_rev

    ax2.set_xticks(x_idx)
    ax2.set_xticklabels(labels_dates, rotation=35, ha='right', fontsize=8)
    ax2.set_title("CA previsionnel par produit (Top 6)", fontsize=11, fontweight='bold', pad=8)
    ax2.set_ylabel("CA previsionnel (MAD)", fontsize=9)
    ax2.legend(fontsize=7, loc='upper left', ncol=2)
    ax2.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f'{x:,.0f} MAD'))
    ax2.grid(True, axis='y', linestyle=':', alpha=0.5)

    # --- PANNEAU 3 : TOP 12 PRODUITS EN VOLUME (prochain mois) ---
    ax3 = fig.add_subplot(2, 2, 3)
    col_qty_p3 = 'Qty_Prev_Prophet' if (config.ACTIVER_PROPHET and 'Qty_Prev_Prophet' in list(dict_prevision_prod.values())[0].columns) \
                 else 'Qty_Prev_Decompo_Saisonniere'
    prochaine_date = df_prev_total['Date'].iloc[0]
    date_str = pd.to_datetime(prochaine_date).strftime("%B %Y")

    volumes_prod = {}
    for prod in produits_uniques:
        df_fc_p = dict_prevision_prod[prod]
        if col_qty_p3 in df_fc_p.columns and len(df_fc_p) > 0:
            volumes_prod[prod] = float(df_fc_p[col_qty_p3].iloc[0])
    top12_vol = sorted(volumes_prod, key=volumes_prod.get, reverse=True)[:12]
    noms = [p[:28] for p in top12_vol]
    vals_vol = [volumes_prod[p] for p in top12_vol]

    couleurs_barres = ['#e74c3c' if v > np.median(vals_vol) else '#3498db' for v in vals_vol]
    bars = ax3.barh(noms[::-1], vals_vol[::-1], color=couleurs_barres[::-1], edgecolor='white', height=0.65)
    for bar, v in zip(bars, vals_vol[::-1]):
        ax3.text(bar.get_width() * 1.01, bar.get_y() + bar.get_height()/2,
                 f" {v:,.0f}", va='center', ha='left', fontsize=8, fontweight='bold', color='#2c3e50')

    ax3.set_title(f"Top 12 produits - Volume prévu {date_str}", fontsize=11, fontweight='bold', pad=8)
    ax3.set_xlabel("Quantité prévue (unités)", fontsize=9)
    ax3.set_xlim(0, max(vals_vol) * 1.25 if vals_vol else 1)
    ax3.tick_params(axis='y', labelsize=8)
    ax3.grid(True, axis='x', linestyle=':', alpha=0.5)

    # --- PANNEAU 4 : TOP 15 PRODUITS PAR CA PRÉVU (prochain mois) ---
    ax4 = fig.add_subplot(2, 2, 4)
    col_rev_p4 = 'Rev_Prev_Prophet' if (config.ACTIVER_PROPHET and 'Rev_Prev_Prophet' in list(dict_prevision_prod.values())[0].columns) \
                 else 'Rev_Prev_Decompo_Saisonniere'

    ca_prod_p4 = {}
    for prod in produits_uniques:
        df_fc_p4 = dict_prevision_prod[prod]
        if col_rev_p4 in df_fc_p4.columns and len(df_fc_p4) > 0:
            ca_prod_p4[prod] = float(df_fc_p4[col_rev_p4].iloc[0])

    top15_ca = sorted(ca_prod_p4, key=ca_prod_p4.get, reverse=True)[:15]
    noms_ca  = [p[:30] for p in top15_ca]
    vals_ca  = [ca_prod_p4[p] for p in top15_ca]

    coul_ca = ['#27ae60' if v > np.median(vals_ca) else '#2ecc71' for v in vals_ca]
    bars4 = ax4.barh(noms_ca[::-1], vals_ca[::-1], color=coul_ca[::-1], edgecolor='white', height=0.65)
    for bar, v in zip(bars4, vals_ca[::-1]):
        ax4.text(bar.get_width() * 1.01, bar.get_y() + bar.get_height() / 2,
                 f" {v:,.0f} MAD", va='center', ha='left', fontsize=7.5, fontweight='bold', color='#2c3e50')

    ax4.set_title(f"Top 15 produits - CA prévu {date_str}", fontsize=11, fontweight='bold', pad=8)
    ax4.set_xlabel("CA prévisionnel (MAD)", fontsize=9)
    ax4.set_xlim(0, max(vals_ca) * 1.30 if vals_ca else 1)
    ax4.tick_params(axis='y', labelsize=8)
    ax4.grid(True, axis='x', linestyle=':', alpha=0.5)
    ax4.xaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f'{x:,.0f}'))

    plt.suptitle(
        f"TABLEAU DE BORD GERANT | BOULANGERIE PAUL | Previsions sur {config.FORECAST_PERIODS} mois",
        fontsize=14, fontweight='bold', y=1.01, color='#2c3e50'
    )
    plt.tight_layout(rect=[0, 0, 1, 0.99])

    os.makedirs(output_dir, exist_ok=True)
    graph_path = os.path.join(output_dir, "dashboard_gerant.png")
    plt.savefig(graph_path, dpi=150, bbox_inches='tight', facecolor='#f8f9fa')
    plt.close(fig)
    logger.info("Tableau de bord gérant sauvegardé dans : '%s'", graph_path)


# ==============================================================================
# DASHBOARD MATIÈRES PREMIÈRES (BOM / MRP)
# ==============================================================================
def tracer_dashboard_matieres_premieres(df_besoins_mrp, output_dir):
    """Dashboard centré sur le mois suivant : top 20 ingrédients, tendance, familles."""
    if df_besoins_mrp is None or df_besoins_mrp.empty:
        logger.info("Aucune donnée matières premières à afficher (BOM vide).")
        return

    COULEURS = [
        '#2ecc71', '#3498db', '#e74c3c', '#f39c12', '#9b59b6',
        '#1abc9c', '#e67e22', '#c0392b', '#2980b9', '#27ae60',
        '#d35400', '#8e44ad', '#16a085', '#f1c40f', '#95a5a6',
        '#bdc3c7', '#7f8c8d', '#2c3e50', '#e8daef', '#d5e8d4'
    ]

    def extraire_unite(nom):
        n = str(nom)
        m = re.search(r'\(([^)]+)\)\s*$', n)
        if m:
            return m.group(1)
        n_maj = n.upper()
        if any(k in n_maj for k in [
            "FARINE", "SEMOULE", "BEURRE", "SUCRE", "SEL", "LEVURE",
            "ŒUF", "OEU", "CHOCOLAT", "POUDRE", "CRÈME", "CREME",
            "PÂTE", "COMPOTE", "RAISINS", "FRUITS", "MENTHE",
            "GLAÇONS", "GLACE", "GANACHE", "FONDANT", "GELÉE",
            "AMANDE", "MIEL", "CHAPELURE", "CITRON", "MANGUE",
            "GARNITURE", "SALADE", "JAMBON", "POULET", "SAUMON"
        ]):
            return "g"
        if any(k in n_maj for k in [
            "LAIT", "JUS", "SIROP", "ARÔME", "AROME", "SAUCE", "EAU", "CAFÉ", "COFFEE"
        ]):
            return "ml"
        return "unité"

    def nom_court(nom):
        return re.sub(r'\s*\([^)]*\)\s*$', '', str(nom))

    def famille_ingredient(nom):
        n = str(nom).upper()
        if any(k in n for k in ["FARINE", "SEMOULE", "LEVURE", "SEL"]):
            return "Farines & Levures"
        if any(k in n for k in ["BEURRE", "HUILE", "CRÈME", "CREME", "LAIT"]):
            return "Matières grasses & Lait"
        if any(k in n for k in ["SUCRE", "MIEL", "CONFITURE"]):
            return "Sucrants"
        if any(k in n for k in ["ŒUF", "OEU", "BLANC D"]):
            return "Œufs"
        if any(k in n for k in ["CHOCOLAT", "CACAO", "PÉPITE", "PEPITE"]):
            return "Chocolat"
        if any(k in n for k in ["CAFÉ", "CAFE", "CAPSULE"]):
            return "Café"
        if any(k in n for k in ["SACHET", "MENTHE", "SIROP"]):
            return "Thés & Sirops"
        if any(k in n for k in ["FRUITS", "POMME", "CITRON", "MANGUE", "RAISIN", "JUS"]):
            return "Fruits & Jus"
        if any(k in n for k in ["GOBELET", "EMBALLAGE", "BOÎTE", "BOITE"]):
            return "Emballages"
        if any(k in n for k in ["JAMBON", "POULET", "SAUMON", "POISSON"]):
            return "Protéines"
        if any(k in n for k in ["FROMAGE", "SALADE", "GARNITURE", "SAUCE"]):
            return "Traiteur"
        return "Autres"

    dates_triees = sorted(df_besoins_mrp['Date'].unique())
    mois_suivant = dates_triees[0]
    mois_str     = pd.Timestamp(mois_suivant).strftime("%B %Y")

    df_mois = (
        df_besoins_mrp[df_besoins_mrp['Date'] == mois_suivant]
        .groupby('Ingredient')['Quantite_Requise'].sum()
        .sort_values(ascending=False).reset_index()
    )
    df_mois['Unite']    = df_mois['Ingredient'].apply(extraire_unite)
    df_mois['NomCourt'] = df_mois['Ingredient'].apply(nom_court)

    fig = plt.figure(figsize=(18, 13))
    fig.patch.set_facecolor('#f8f9fa')

    # PANNEAU 1 : Bon de commande mois suivant
    ax1 = fig.add_subplot(2, 2, (1, 2))
    top20     = df_mois.head(20).copy()
    top20_inv = top20.iloc[::-1]
    barres = ax1.barh(top20_inv['NomCourt'], top20_inv['Quantite_Requise'],
                      color=[COULEURS[i % len(COULEURS)] for i in range(len(top20_inv))],
                      edgecolor='white', height=0.72)
    for bar, (_, row) in zip(barres, top20_inv.iterrows()):
        v = row['Quantite_Requise']
        ax1.text(bar.get_width() * 1.007, bar.get_y() + bar.get_height() / 2,
                 f"{v:,.0f} {row['Unite']}", va='center', ha='left',
                 fontsize=8.5, fontweight='bold', color='#2c3e50')

    ax1.set_title(f"Bon de commande — Matières premières à approvisionner : {mois_str}",
                  fontsize=12, fontweight='bold', pad=10)
    ax1.set_xlabel("Quantité à commander (unité selon ingrédient)", fontsize=9)
    ax1.set_xlim(0, top20_inv['Quantite_Requise'].max() * 1.25 if len(top20_inv) else 1)
    ax1.tick_params(axis='y', labelsize=9)
    ax1.grid(True, axis='x', linestyle=':', alpha=0.5)
    ax1.xaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f'{x:,.0f}'))

    # PANNEAU 2 : Tendance 12 mois des 6 ingrédients principaux
    ax2 = fig.add_subplot(2, 2, 3)
    top6_noms = df_mois.head(6)['Ingredient'].tolist()
    df_pivot = df_besoins_mrp.pivot_table(index='Date', columns='Ingredient',
                                          values='Quantite_Requise', aggfunc='sum').fillna(0).sort_index()
    dates_str = [pd.Timestamp(d).strftime('%b %Y') for d in df_pivot.index]
    x_idx = range(len(dates_str))

    for i, ing in enumerate(top6_noms):
        if ing not in df_pivot.columns:
            continue
        vals = df_pivot[ing].values
        ax2.plot(x_idx, vals, marker='o', linewidth=2, markersize=5,
                 color=COULEURS[i], label=nom_court(ing)[:28])
        ax2.fill_between(x_idx, vals, alpha=0.07, color=COULEURS[i])

    ax2.axvline(x=0, color='#e74c3c', linewidth=1.5, linestyle='--', alpha=0.7,
                label=f'← {mois_str} (commande)')
    ax2.set_xticks(list(x_idx))
    ax2.set_xticklabels(dates_str, rotation=35, ha='right', fontsize=8)
    ax2.set_title("Tendance sur 12 mois — Top 6 ingrédients", fontsize=11, fontweight='bold', pad=8)
    ax2.set_ylabel("Quantité requise", fontsize=9)
    ax2.legend(fontsize=7.5, loc='upper left', ncol=1)
    ax2.grid(True, linestyle=':', alpha=0.5)
    ax2.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f'{x:,.0f}'))

    # PANNEAU 3 : Camembert familles mois suivant
    ax3 = fig.add_subplot(2, 2, 4)
    df_fam_mois = df_mois.copy()
    df_fam_mois['Famille'] = df_fam_mois['Ingredient'].apply(famille_ingredient)
    df_familles = df_fam_mois.groupby('Famille')['Quantite_Requise'].sum().sort_values(ascending=False)

    total_fam = df_familles.sum()
    masque_petit = (df_familles / total_fam) < 0.02
    autres_val   = df_familles[masque_petit].sum()
    df_familles  = df_familles[~masque_petit].copy()
    if autres_val > 0:
        df_familles['Autres'] = autres_val

    wedges, texts, autotexts = ax3.pie(
        df_familles.values, labels=df_familles.index, autopct='%1.1f%%',
        colors=[COULEURS[i % len(COULEURS)] for i in range(len(df_familles))],
        startangle=90, pctdistance=0.75,
        wedgeprops=dict(linewidth=1.2, edgecolor='white')
    )
    for t in texts:
        t.set_fontsize(8.5)
    for at in autotexts:
        at.set_fontsize(7.5)
        at.set_fontweight('bold')

    ax3.set_title(f"Répartition par famille — {mois_str}", fontsize=11, fontweight='bold', pad=8)

    plt.suptitle(f"APPROVISIONNEMENT MATIÈRES PREMIÈRES | PAUL CASABLANCA | {mois_str}",
                 fontsize=13, fontweight='bold', y=1.01, color='#2c3e50')
    plt.tight_layout(rect=[0, 0, 1, 0.99])
    os.makedirs(output_dir, exist_ok=True)
    path_mp = os.path.join(output_dir, "dashboard_matieres_premieres.png")
    plt.savefig(path_mp, dpi=150, bbox_inches='tight', facecolor='#f8f9fa')
    plt.close(fig)
    logger.info("Dashboard matières premières sauvegardé : '%s'", path_mp)


# ==============================================================================
# EXPORT EXCEL MULTI-FEUILLES
# ==============================================================================
def exporter_excel_previsions(df_hist_total, df_prev_total, dict_prevision_prod,
                              produits_uniques, output_dir, df_besoins_mrp=None):
    """Génère un fichier Excel multi-feuilles avec mise en forme professionnelle."""
    try:
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        logger.warning("openpyxl non installé — export Excel ignoré (pip install openpyxl).")
        return

    chemin_xl = os.path.join(output_dir, "previsions_PAUL_2026.xlsx")

    H_FILL   = PatternFill("solid", fgColor="2C3E50")
    H_FONT   = Font(color="FFFFFF", bold=True, size=10)
    ALT_FILL = PatternFill("solid", fgColor="EBF5FB")
    GRN_FILL = PatternFill("solid", fgColor="D5F5E3")
    ORG_FILL = PatternFill("solid", fgColor="FDEBD0")
    CENTER   = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT     = Alignment(horizontal="left", vertical="center")
    thin     = Side(border_style="thin", color="CCCCCC")
    BORDER   = Border(left=thin, right=thin, top=thin, bottom=thin)

    def styler(ws, titre, col_premiere_large=True):
        max_col = ws.max_column
        ws.insert_rows(1)
        ws['A1'] = titre
        ws['A1'].font = Font(bold=True, size=13, color="2C3E50")
        ws['A1'].alignment = CENTER
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
        ws.row_dimensions[1].height = 28
        for cell in ws[2]:
            cell.fill = H_FILL
            cell.font = H_FONT
            cell.alignment = CENTER
            cell.border = BORDER
        for r_idx, row in enumerate(ws.iter_rows(min_row=3, max_row=ws.max_row), 1):
            for cell in row:
                if r_idx % 2 == 0:
                    cell.fill = ALT_FILL
                cell.alignment = CENTER
                cell.border = BORDER
            if col_premiere_large:
                row[0].alignment = LEFT
        for col in ws.columns:
            try:
                max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col)
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 3, 24)
            except Exception:
                pass

    mois_labels = [pd.to_datetime(d).strftime('%b %Y') for d in df_prev_total['Date'].values]
    col_hw_rev  = 'Rev_Prev_Holt_Winters'
    col_hw_qty  = 'Qty_Prev_Holt_Winters'

    with pd.ExcelWriter(chemin_xl, engine='openpyxl') as writer:
        # Feuille 1
        cols_f1 = {
            'Date': 'Mois',
            'Rev_Prev_Holt_Winters': 'CA Holt-Winters (MAD)',
            'Qty_Prev_Holt_Winters': 'QT Holt-Winters',
            'Rev_Prev_Decompo_Saisonniere': 'CA Décompo (MAD)',
            'Rev_Prev_Moyenne_Mobile': 'CA Moy. Mobile (MAD)',
            'Rev_Prev_Tendance': 'CA Tendance (MAD)',
        }
        cols_existantes = [c for c in cols_f1 if c in df_prev_total.columns]
        df_f1 = df_prev_total[cols_existantes].rename(columns=cols_f1).copy()
        df_f1['Mois'] = pd.to_datetime(df_f1['Mois']).dt.strftime('%B %Y')
        for c in df_f1.columns[1:]:
            df_f1[c] = df_f1[c].round(0).astype(int)
        df_f1.to_excel(writer, sheet_name='Prévisions Globales', index=False, startrow=1)

        # Feuille 2
        lignes = []
        for prod in produits_uniques:
            df_p = dict_prevision_prod[prod]
            if col_hw_rev not in df_p.columns:
                continue
            row = {'Produit': prod}
            total_ca = total_qt = 0.0
            for i, label in enumerate(mois_labels):
                ca = float(df_p[col_hw_rev].iloc[i]) if i < len(df_p) else 0.0
                qt = float(df_p[col_hw_qty].iloc[i]) if (col_hw_qty in df_p.columns and i < len(df_p)) else 0.0
                row[f'CA {label} (MAD)'] = round(ca, 0)
                row[f'QT {label}'] = round(qt, 0)
                total_ca += ca
                total_qt += qt
            row['TOTAL CA (MAD)'] = round(total_ca, 0)
            row['TOTAL QT'] = round(total_qt, 0)
            lignes.append(row)

        df_f2 = pd.DataFrame(lignes).sort_values('TOTAL CA (MAD)', ascending=False)
        df_f2.to_excel(writer, sheet_name='Par Produit', index=False, startrow=1)

        # Feuille 3
        df_f3 = df_f2[['Produit', 'TOTAL CA (MAD)', 'TOTAL QT']].head(30).copy()
        df_f3.insert(0, 'Rang', range(1, len(df_f3) + 1))
        df_f3.to_excel(writer, sheet_name='Top 30 Produits', index=False, startrow=1)

        # Feuille 4
        df_f4 = df_hist_total[['Date', 'Chiffre_Affaires_Total', 'Quantite_Total']].copy()
        df_f4['Date'] = pd.to_datetime(df_f4['Date']).dt.strftime('%B %Y')
        df_f4.columns = ['Mois', 'CA Réel (MAD)', 'Quantité Réelle']
        df_f4['CA Réel (MAD)'] = df_f4['CA Réel (MAD)'].round(0).astype(int)
        df_f4['Quantité Réelle'] = df_f4['Quantité Réelle'].round(0).astype(int)
        df_f4.to_excel(writer, sheet_name='Historique', index=False, startrow=1)

        # Feuille 5
        mois_label = None
        if df_besoins_mrp is not None and not df_besoins_mrp.empty:
            def _extraire_unite_xl(nom):
                m = re.search(r'\(([^)]+)\)\s*$', str(nom))
                if m:
                    return m.group(1)
                n = str(nom).upper()
                if any(k in n for k in ["FARINE", "SEMOULE", "BEURRE", "SUCRE", "SEL",
                                        "LEVURE", "ŒUF", "OEU", "CHOCOLAT", "POUDRE",
                                        "CRÈME", "CREME", "PÂTE", "COMPOTE", "RAISINS",
                                        "FRUITS", "MENTHE", "GLAÇONS", "GLACE", "GANACHE",
                                        "FONDANT", "GELÉE", "AMANDE", "MIEL", "CHAPELURE",
                                        "CITRON", "MANGUE", "GARNITURE", "SALADE",
                                        "JAMBON", "POULET", "SAUMON"]):
                    return "g"
                if any(k in n for k in ["LAIT", "JUS", "SIROP", "ARÔME", "AROME", "SAUCE"]):
                    return "ml"
                return "unité"

            def _nom_court_xl(nom):
                return re.sub(r'\s*\([^)]*\)\s*$', '', str(nom))

            mois_suivant = sorted(df_besoins_mrp['Date'].unique())[0]
            mois_label   = pd.Timestamp(mois_suivant).strftime("%B %Y")

            df_bdc = (
                df_besoins_mrp[df_besoins_mrp['Date'] == mois_suivant]
                .groupby('Ingredient')['Quantite_Requise'].sum()
                .sort_values(ascending=False).reset_index()
            )
            df_bdc['Ingrédient']      = df_bdc['Ingredient'].apply(_nom_court_xl)
            df_bdc['Unité']           = df_bdc['Ingredient'].apply(_extraire_unite_xl)
            df_bdc['Qté à commander'] = df_bdc['Quantite_Requise'].round(1)
            df_bdc['Qté (kg ou L)']   = df_bdc.apply(
                lambda r: round(r['Quantite_Requise'] / 1000, 2)
                          if r['Unité'] in ('g', 'ml') else r['Quantite_Requise'], axis=1)
            df_bdc['Fournisseur']  = ""
            df_bdc['Commande OK?'] = ""
            cols_bdc = ['Ingrédient', 'Unité', 'Qté à commander', 'Qté (kg ou L)',
                        'Fournisseur', 'Commande OK?']
            df_bdc[cols_bdc].to_excel(writer, sheet_name='Bon de Commande', index=False, startrow=1)

        # Mise en forme
        wb = writer.book
        styler(wb['Prévisions Globales'], "Prévisions Globales PAUL – 2026 (modèle Holt-Winters)")
        styler(wb['Par Produit'], "Prévisions CA et Quantités par produit – 2026")
        styler(wb['Top 30 Produits'], "Top 30 produits par CA prévisionnel total – 2026")
        styler(wb['Historique'], "Historique CA mensuel réel – 2021 à 2025")

        if 'Bon de Commande' in wb.sheetnames:
            styler(wb['Bon de Commande'], f"Bon de commande matières premières – {mois_label}",
                   col_premiere_large=True)
            bdc_ws = wb['Bon de Commande']
            RED_FILL = PatternFill("solid", fgColor="7B241C")
            for cell in bdc_ws[2]:
                cell.fill = RED_FILL
                cell.font = Font(color="FFFFFF", bold=True, size=10)
            bdc_ws.column_dimensions['F'].width = 14

        top_ws = wb['Top 30 Produits']
        for cell in top_ws[3]:
            cell.fill = GRN_FILL
            cell.font = Font(bold=True)

        glob_ws = wb['Prévisions Globales']
        for cell in glob_ws[2]:
            if 'Holt' in str(cell.value):
                col_letter = get_column_letter(cell.column)
                for r in range(2, glob_ws.max_row + 1):
                    glob_ws[f'{col_letter}{r}'].fill = ORG_FILL
                break

    logger.info("Export Excel prévisions → '%s'", chemin_xl)
