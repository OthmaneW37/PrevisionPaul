# -*- coding: utf-8 -*-
"""
Importe le classeur « docs/Recettes PAUL_Cuisinep2.xlsx » (2e lot de fiches chef)
dans data/recettes_exactes.json (+ provenance).

Format de CE classeur (différent des précédents) : une feuille « Ingrédients »
= une GRANDE TABLE À PLAT, une ligne par (recette, ingrédient), colonnes :
  Recette | Code | Ingrédient | Quantité | Unité | Poids (Kg/g)

Conversion des quantités : on s'appuie sur la colonne « Poids (Kg/g) » (poids déjà
résolu, y compris pour les œufs/pièces). Son échelle est kg OU g selon la recette
-> on détecte : si la recette utilise l'unité « KG » quelque part, ses poids sont
en kg (×1000), sinon en g. (Les « càS », « branche »… sont déjà résolus dans cette
colonne, donc pas de conversion spéciale à faire.)

Chaque recette est soit un PRODUIT VENDU (rattaché à son nom exact dans les ventes
via PLAN, validé avec le gérant), soit une SOUS-PRÉPARATION (base/sauce) stockée
sous un nom canonique et réutilisable comme ingrédient.

Décidé avec le gérant (juillet 2026) :
  - Croissant Saumon Bénédicte -> BRIOCHE SAUMON BENEDICTE
  - Poulet Champignon Moutarde -> SUPREME POULET SCE CHAMPINION

Lancement :
  python outils/importer_recettes_cuisine_p2.py --dry-run
  python outils/importer_recettes_cuisine_p2.py
"""

# --- Script utilitaire : exécutable depuis n'importe où (se cale sur la racine) ---
import os as _os, sys as _sys
_RACINE = _os.path.dirname(_os.path.dirname(_os.path.abspath(__file__)))
_sys.path.insert(0, _RACINE)
_os.chdir(_RACINE)
# ---------------------------------------------------------------------------------

import argparse
import json
import os
import shutil
import unicodedata
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook

XLSX = os.path.join(_RACINE, "docs", "Recettes PAUL_Cuisinep2.xlsx")
RECETTES_JSON = os.path.join(_RACINE, "data", "recettes_exactes.json")
PROVENANCE_JSON = os.path.join(_RACINE, "data", "recettes_exactes_provenance.json")
VENTES = os.path.join(_RACINE, "donnees_ventes", "ventes_journalieres.csv")

# Recette (feuille) -> produit(s) vendu(s), résolus ensuite au nom EXACT des ventes.
PLAN_VENDUS = {
    "omelettecomplete":               ["Omelette Complete"],
    "hambourgeoisloriginal":          ["Hambourgeois Original"],
    "briochepouletavocat":            ["Brioche Poulet Avocat"],
    "hambourgeoislesignature":        ["Hambourgeois Le Signature"],
    "minihambourgeoispouletpane":     ["Mini Hambourgeois"],
    "emincedeboeufsaucechampignon":   ["Emince Boeuf Sauce Champignon"],
    "pavedesaumonsauceagrumes":       ["Pave saumon sauce agrumes", "Pave saumon sauce agrume PM"],
    "filetdeboeufsaucemaison":        ["Filet Boeuf Sauce Maison"],
    "poulet5grainesfaconcordonbleu":  ["Poulet 5 grain cordon bleu"],
    "omelettesouffleetruffeparmesan": ["Omelette Soufflee Truffe Parmesan"],
    "couscouspouletplat":             ["Couscous poulet"],
    "couscousboeufoupouletplat":      ["Couscous Boeuf"],
    "oeufsbrouillesalatruffe":        ["Oeufs brouill truffe"],
    "croissantlemediterraneen":       ["Croissant Mediterraneen"],
    "croissantvegetarien":            ["Croissant Sale Vegetarien"],
    "omelettesouffleeforestiere":     ["Omelette Soufflee Forestiere"],
    "pouletchampignonmoutarde":       ["Supreme poulet sce champinion"],   # décision gérant
    "briochelemediterraneen":         ["Brioche Mediterraneen"],
    "croissantsaumonbenedicte":       ["Brioche saumon benedicte"],        # décision gérant
    "saucetigrequipleure":            ["Tigre qui pleure"],
    "carottesroties":                 ["Accompagnement carottes roties"],
    # --- 3e vague (juillet 2026) ---
    "minicroissantpdjenfant":         ["Mini croiss PDJ enfant"],
    "croissantjambonfromage":         ["Croissant jambon from"],
    "swthontresse":                   ["SW tresse au thon"],
    "swpouletpanevae":                ["SW poulet pane VSP"],
    "miniquichelorrainevae":          ["Mini quiche lorraine"],
    "miniquichesaumonepinardsvae":    ["Mini quiche saumon epinards"],
    "sandwichtomatemozzapesto":       ["Sandwich tomate mozzarella"],
    "granolapassionframboise":        ["From BL granola passion framboise"],
    "saladecaesarvsp":                ["Salade caesar"],
    "sandwichledieppois":             ["Sandwich dieppois"],
    "sandwichatlantiquevae":          ["SW atlantique (VSP)"],
    "sandwichmontagnardvae":          ["SW montagnard (VSP)"],
    "sandwichhotdog":                 ["SW hot dog", "Navette hot dog"],   # décision gérant
    "chiapuddingfruitrouge":          ["Chia pudding pass framb"],         # décision gérant
    "fromageblancgranola":            ["From BL granola miel"],            # décision gérant
    "fromageblancnature":             ["From BL nature VSP"],
    "jaouharasurplace":               ["Jawhara pistache"],                # décision gérant
    "doubletartinepastrami":          ["Double tartine pastrami"],
}

# Recette -> sous-préparation (nom canonique en majuscules), pas un produit vendu.
PLAN_SOUS = {
    "melangeavocat":       "MELANGE AVOCAT",
    "viandehacheemaison":  "VIANDE HACHEE MAISON",
    "saucegrecque":        "SAUCE GRECQUE",
    "sauceasia":           "SAUCE ASIA",
    "chakchouka":          "CHAKCHOUKA",
    "basecouscous":        "BASE COUSCOUS",
    "rizalaciboulette":    "RIZ A LA CIBOULETTE",
    "crevettegrisepanee":  "CREVETTE GRISE PANEE",
    "pureealatruffe":      "PUREE A LA TRUFFE",
    # --- 3e vague (juillet 2026) : sauces/bases des burgers & sandwichs ---
    "saucetomate":         "SAUCE TOMATE",
    "oignonscaramelises":  "OIGNONS CARAMELISES",
    "oignonsfrits":        "OIGNONS FRITS",
    "saucecheddar":        "SAUCE CHEDDAR",
    "saucehollandaise":    "SAUCE HOLLANDAISE",
    "appareilajaouhara":   "APPAREIL A JAOUHARA",
}


def _norm(s):
    s = str(s or "").lower().replace("œ", "oe").replace("æ", "ae")
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    return "".join(c for c in s if c.isalnum())   # ignore apostrophes/espaces/ponctuation


def _index_ventes():
    df = pd.read_csv(VENTES, sep=";", encoding="utf-8")
    idx = {}
    for nom in df["Produit"].dropna().astype(str).unique():
        idx.setdefault(_norm(nom), nom)
    return idx


def lire_recettes():
    """Lit la feuille Ingrédients -> {recette: {ingrédient (g|unité): qté}}."""
    wb = load_workbook(XLSX, data_only=True)
    feuille = next(s for s in wb.sheetnames if _norm(s) == "ingredients")
    ws = wb[feuille]

    lignes = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row or not row[0] or not row[2]:
            continue
        recette = str(row[0]).strip()
        ingr = str(row[2]).strip()
        qte = row[3]
        unite = str(row[4] or "").strip().upper()
        poids = row[5] if len(row) > 5 else None
        lignes.append((recette, ingr, qte, unite, poids))

    # échelle par recette : kg si l'unité « KG » y apparaît, sinon g
    from collections import defaultdict
    unites = defaultdict(set)
    for rec, _, _, u, _ in lignes:
        unites[rec].add(u)

    recettes = {}
    for rec, ingr, qte, unite, poids in lignes:
        echelle_kg = "KG" in unites[rec]
        val = None
        if poids is not None and str(poids).strip() not in ("", "None"):
            try:
                val = float(str(poids).replace(",", ".")) * (1000 if echelle_kg else 1)
                base = "g"
            except ValueError:
                val = None
        if val is None:                      # repli : quantité + unité
            try:
                q = float(str(qte).replace(",", "."))
            except (ValueError, TypeError):
                continue
            if unite == "KG":
                val, base = q * 1000, "g"
            elif unite in ("G",):
                val, base = q, "g"
            else:                            # U / UNITE(S) / branche…
                val, base = q, "unité"
        recettes.setdefault(rec, {})[f"{ingr} ({base})"] = round(val, 2)
    return recettes


def construire():
    """-> (recettes {nom_cible: recette}, notes, anomalies, non_mappes)."""
    brutes = lire_recettes()
    idx = _index_ventes()
    with open(RECETTES_JSON, encoding="utf-8") as f:
        deja = {_norm(k) for k in json.load(f)}

    sorties, notes, anomalies, non_mappes = {}, [], [], []
    for rec, contenu in brutes.items():
        rn = _norm(rec)
        if rn in PLAN_VENDUS:
            for approx in PLAN_VENDUS[rn]:
                nom_exact = idx.get(_norm(approx))
                if nom_exact is None:
                    anomalies.append(f"« {rec} » : cible « {approx} » absente des ventes -> ignorée.")
                    continue
                sorties[nom_exact] = dict(contenu)
                notes.append(f"« {rec} » -> {nom_exact} ({len(contenu)} ingr).")
        elif rn in PLAN_SOUS:
            canon = PLAN_SOUS[rn]
            sorties[canon] = dict(contenu)
            notes.append(f"« {rec} » -> sous-recette {canon} ({len(contenu)} ingr).")
        else:
            non_mappes.append(rec)
    return sorties, notes, anomalies, non_mappes


def appliquer(recettes):
    with open(RECETTES_JSON, encoding="utf-8") as f:
        existantes = json.load(f)
    sauvegarde = f"{RECETTES_JSON}.bak-{datetime.now():%Y%m%d-%H%M%S}"
    shutil.copyfile(RECETTES_JSON, sauvegarde)
    existantes.update(recettes)
    with open(RECETTES_JSON, "w", encoding="utf-8") as f:
        json.dump(existantes, f, ensure_ascii=False, indent=2)

    try:
        with open(PROVENANCE_JSON, encoding="utf-8") as f:
            prov = json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        prov = {}
    for prod in recettes:
        prov[prod] = {"source": "fiches chef cuisine (xlsx p2)",
                      "fiche": os.path.basename(XLSX), "score": 1.0,
                      "date": f"{datetime.now():%Y-%m-%d}"}
    with open(PROVENANCE_JSON, "w", encoding="utf-8") as f:
        json.dump(prov, f, ensure_ascii=False, indent=2)
    return sauvegarde


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()
    recettes, notes, anomalies, non_mappes = construire()

    print("=" * 66)
    print(f"  IMPORT RECETTES CUISINE p2{'   [DRY-RUN]' if args.dry_run else ''}")
    print("=" * 66)
    for prod, ings in recettes.items():
        print(f"\n> {prod}  ({len(ings)} ingrédient(s))")
        for ing, q in ings.items():
            print(f"    - {ing:<46} {q}")
    if notes:
        print("\nCorrespondances :")
        for n in notes:
            print(f"  - {n}")
    if anomalies:
        print("\nAnomalies :")
        for a in anomalies:
            print(f"  ! {a}")
    if non_mappes:
        print("\nNon importées (pas de produit vendu correspondant ; à saisir/rattacher) :")
        for r in non_mappes:
            print(f"  - {r}")

    if args.dry_run:
        print("\n[DRY-RUN] Rien écrit.")
        return
    if not recettes:
        print("\nRien à importer.")
        return
    sauvegarde = appliquer(recettes)
    print(f"\n[OK] {len(recettes)} entrée(s) écrite(s). Sauvegarde : {os.path.basename(sauvegarde)}")
    print("     -> Relancer python main.py.")


if __name__ == "__main__":
    main()
