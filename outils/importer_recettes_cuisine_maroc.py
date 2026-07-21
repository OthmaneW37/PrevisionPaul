# -*- coding: utf-8 -*-
"""
Importe les fiches techniques chef « docs/Recettes_PAUL_Maroc_Cuisine.xlsx »
(catégorie CUISINE) dans data/recettes_exactes.json (+ provenance).

Ce classeur a un format DIFFÉRENT du gabarit officiel
(recettes_produits_a_completer.xlsx) : UNE FEUILLE PAR RECETTE, avec les
ingrédients en colonne A, la quantité en colonne C et l'unité en colonne D.
Il ne peut donc pas passer par outils/importer_recettes_chefs.py — d'où ce
script dédié, qui :

  - lit chaque fiche (une feuille = une recette) ;
  - normalise les unités en base g / ml / unité (kg->g, cl->ml, « g (kg) »->g…) ;
  - retire les CONSOMMABLES (bol, couteau, étiquette, sachet…) qui ne sont pas
    des matières premières ;
  - corrige quelques erreurs de saisie connues (unité « unité » là où c'est des
    grammes) et signale les quantités suspectes sans les modifier en douce ;
  - isole la « Sauce miel » de la salade exotique (préparation en bulk 72H) dans
    une sous-recette à part, car sa dose par portion n'est pas définie ;
  - fait CORRESPONDRE chaque fiche au(x) vrai(s) nom(s) de produit vendu
    (PLAN ci-dessous, validé avec le gérant) ;
  - fusionne dans recettes_exactes.json (sauvegarde horodatée + provenance).

Décidé avec le gérant (juillet 2026) :
  - Crêpe Norvégienne      -> CREPE SAUMON ATLQ
  - Tartine Poulet Guacamole -> TARTINE POULET AVOCAT
  - Salade Niçoise GM      -> SALADE NICOISE **et** SALADE NICOISE PM
    (PM reprend provisoirement les quantités du grand modèle — à affiner).

Lancement :
  python outils/importer_recettes_cuisine_maroc.py --dry-run   # aperçu, n'écrit rien
  python outils/importer_recettes_cuisine_maroc.py             # applique (avec sauvegarde)

Après import : relancer le calcul (python main.py) pour répercuter sur le bon
de commande matières.
"""

# --- Script utilitaire : exécutable depuis n'importe où (se cale sur la racine) ---
import os as _os, sys as _sys
_RACINE = _os.path.dirname(_os.path.dirname(_os.path.abspath(__file__)))
_sys.path.insert(0, _RACINE)
_os.chdir(_RACINE)
# ---------------------------------------------------------------------------------

import argparse
import json
import os
import re
import shutil
import unicodedata
from datetime import datetime

from openpyxl import load_workbook

XLSX = os.path.join(_RACINE, "docs", "Recettes_PAUL_Cuisine.xlsx")
RECETTES_JSON = os.path.join(_RACINE, "data", "recettes_exactes.json")
PROVENANCE_JSON = os.path.join(_RACINE, "data", "recettes_exactes_provenance.json")

# (facteur vers l'unité de base, unité de base) — identique à importer_recettes_chefs.
CONVERSIONS = {"g": (1, "g"), "kg": (1000, "g"),
               "ml": (1, "ml"), "l": (1000, "ml"), "cl": (10, "ml"),
               "unité": (1, "unité"), "unite": (1, "unité"),
               "pièce": (1, "unité"), "piece": (1, "unité")}

# Fiche (feuille) -> produit(s) vendu(s) cible(s). Les noms cibles doivent être
# EXACTEMENT ceux de ventes_journalieres.csv (majuscules).
PLAN = [
    ("Crêpe Complète",                    ["CREPE COMPLETE"]),
    ("Crêpe Jambon Fromage VAE",          ["CREPE JAM FROM VAE"]),
    ("Crêpe Jambon Fromage VSP",          ["CREPE JAM FROM VSP"]),
    ("Crêpe Poireaux Crevettes",          ["CREPE CREVETTES POIREAUX"]),
    ("Crêpe Norvégienne",                 ["CREPE SAUMON ATLQ"]),
    ("Tartine Poulet Guacamole",          ["TARTINE POULET AVOCAT"]),
    ("Salade Niçoise GM",                 ["SALADE NICOISE", "SALADE NICOISE PM"]),
    ("Salade Exotique Gambas Agrumes",    ["SLD GAMBAS AGRUMES"]),
    ("Crêpe Chocolat Noisette et Noix",   ["CREPE CHOCO NOIX"]),
    ("Crêpe Miel Noix",                   ["CREPE MIEL NOIX"]),
    ("Crêpe Demoiselle Tatin",            ["CREPE DEMOISELLE TATIN"]),
    ("Salade César Gambas",               ["SLD CESAR GMB CROUST"]),
    ("Salade César Poulet Pané",          ["SALADE CESAR POULET PANNE"]),
    ("Salade César Poulet Mariné",        ["SALADE CESAR RMD"]),
    ("Club Sandwich à la Parisienne",     ["CLUB SW SAUMON"]),
]

# Sous-préparations (utilisées comme ingrédients d'autres recettes) : stockées
# sous leur nom canonique en majuscules, pas rattachées à un produit vendu.
SOUS_RECETTES = {
    "Tomates Cerises Rôties":  "TOMATES CERISES ROTIES",
    "Croutons Pain Nordique":  "CROUTONS PAIN NORDIQUE",
    "Appareil Crêpe Sucrée":   "APPAREIL CREPE SUCREE",
    "Sauce à la Moutarde":     "SAUCE MOUTARDE",
    "Side Froid":              "SIDE FROID",
}

# Mots-clés d'un ingrédient à EXCLURE (emballage / couvert / accessoire, pas une
# matière première alimentaire).
CONSOMMABLES = ["bol ", "bolà", "couteau", "fourchette", "cuiller", "cuillère",
                "étiquette", "etiquette", "sachet", "mini pot", "pot cristal",
                "carré rain", "carre rain", "serviette", "barquette", "emballage",
                "sac croissant", "pique ", "vinyl"]

# Corrections de saisie connues, par feuille (nom normalisé, comparé par préfixe
# car Excel tronque les noms d'onglets à 31 caractères) puis par ingrédient :
#   ("unite", "g")  -> l'unité saisie est fausse, forcer cette unité de base ;
#   ("valeur", 2.0) -> la quantité est aberrante, la remplacer par cette valeur (g).
CORRECTIONS = {
    "saladeexotiquegambasagrum": {
        "potiron": ("unite", "g"),     # saisi « unité » alors que ce sont des grammes
        "poivre":  ("valeur", 2.0),    # 70 g de poivre : aberrant -> 2 g
    },
}


def _correction(fiche_norm, ing_norm):
    """Correction éventuelle pour (feuille, ingrédient), tolérante aux noms tronqués."""
    for cle, corr in CORRECTIONS.items():
        if fiche_norm.startswith(cle) or cle.startswith(fiche_norm):
            return corr.get(ing_norm)
    return None


def _norm(s):
    """Minuscule, sans accents ni espaces — pour comparer noms de feuilles/ingrédients."""
    s = unicodedata.normalize("NFKD", str(s or ""))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return "".join(s.lower().split())


def _nombre(v):
    if v is None or (isinstance(v, str) and not v.strip()):
        return None
    try:
        return float(str(v).replace(",", ".").strip())
    except ValueError:
        return None


def convertir(qte, brut):
    """(quantité, unité brute) -> (valeur en base, base 'g'|'ml'|'unité') ou None.

    Gère : « kg » (x1000), « cl » (x10) ; et surtout les unités descriptives
    dont le poids unitaire est donné entre parenthèses — « unité (60g) »,
    « feuille moyenne (16g) », « c.à.s (10g) » -> quantité x ce poids en grammes.
    Le « g (4 unités) » / « g (1U) » reste en grammes (pas de poids en g dans la
    parenthèse). Un poids en g dans la parenthèse est PRIORITAIRE (donne la masse
    réelle même quand le libellé de base est « unité »).
    """
    u = str(brut or "")
    base_token = u.split("(")[0].strip().lower()
    paren = u[u.find("(") + 1:u.rfind(")")] if "(" in u and ")" in u else ""
    m = re.search(r"([\d.,]+)\s*g\b", paren)          # poids unitaire en g ?
    if m:
        grammes = float(m.group(1).replace(",", "."))
        return round(qte * grammes, 2), "g"
    if base_token in CONVERSIONS:
        facteur, base = CONVERSIONS[base_token]
        return round(qte * facteur, 2), base
    return None


def _trouver_feuille(wb, nom_fiche):
    """Retrouve la vraie feuille (les noms Excel sont tronqués à 31 car.)."""
    cible = _norm(nom_fiche)
    for nom in wb.sheetnames:
        n = _norm(nom)
        if n == cible or cible.startswith(n) or n.startswith(cible):
            return nom
    return None


def lire_fiche(ws):
    """Une feuille-recette -> (recette {ingrédient (base): qté}, sauce_miel, anomalies).

    Repère le bloc d'ingrédients (entre l'entête « Ingrédients » et « Poids total »),
    exclut les consommables, applique les corrections d'unité, isole la sauce miel.
    """
    recette, sauce, anomalies = {}, {}, []
    fiche_norm = _norm(ws.title)
    dans_bloc = False
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        col0 = str(row[0]).strip() if row and row[0] else ""
        n0 = _norm(col0)
        if n0.startswith("ingredient"):        # entête du bloc
            dans_bloc = True
            continue
        if n0.startswith("poidstotal"):        # fin du bloc
            break
        if not dans_bloc or not col0:
            continue

        qte = _nombre(row[2] if len(row) > 2 else None)
        unite_brut = row[3] if len(row) > 3 else None
        if qte is None:
            continue

        ing_norm = _norm(col0)
        if any(k in col0.lower() for k in CONSOMMABLES):
            anomalies.append(f"{ws.title} : « {col0} » exclu (consommable/emballage).")
            continue

        # sauce miel de la salade exotique : préparation en bulk -> sous-recette à part
        if col0.lower().startswith("sauce miel"):
            nom_ing = col0.split(":", 1)[-1].strip() or col0
            conv = convertir(qte, unite_brut) or (round(qte, 2), "g")
            val_s, base_s = conv
            sauce[f"{nom_ing} ({base_s})"] = val_s
            continue

        if qte <= 0:
            continue

        # correction de saisie connue (unité fausse, ou quantité aberrante)
        corr = _correction(fiche_norm, ing_norm)
        if corr and corr[0] == "unite":
            facteur, base = CONVERSIONS[corr[1]]
            val = round(qte * facteur, 2)
        elif corr and corr[0] == "valeur":
            val, base = corr[1], "g"
            anomalies.append(f"{ws.title} : « {col0} » corrigé à {val} g "
                             f"(saisie {qte} jugée aberrante).")
        else:
            conv = convertir(qte, unite_brut)
            if conv is None:
                anomalies.append(f"{ws.title} : unité « {unite_brut} » inconnue pour "
                                 f"« {col0} » -> ligne ignorée.")
                continue
            val, base = conv
            # garde-fou : quantité énorme pour un assaisonnement (poivre/sel > 20 g)
            if base == "g" and any(m in ing_norm for m in ("poivre", "sel")) and val > 20:
                anomalies.append(f"{ws.title} : « {col0} » = {val} g semble ÉLEVÉ pour "
                                 f"un assaisonnement (importé tel quel, à vérifier).")
        recette[f"{col0} ({base})"] = val
    return recette, sauce, anomalies


def construire(chemin_xlsx=XLSX):
    """Lit tout le classeur -> (recettes {produit: {ing: qté}}, anomalies, notes)."""
    wb = load_workbook(chemin_xlsx, data_only=True)
    recettes, anomalies, notes = {}, [], []

    # produits vendus
    for nom_fiche, cibles in PLAN:
        feuille = _trouver_feuille(wb, nom_fiche)
        if feuille is None:
            anomalies.append(f"Fiche introuvable dans le classeur : « {nom_fiche} ».")
            continue
        rec, sauce, anos = lire_fiche(wb[feuille])
        anomalies += anos
        if not rec:
            anomalies.append(f"« {nom_fiche} » : aucun ingrédient exploitable -> non importé.")
            continue
        for cible in cibles:
            recettes[cible] = dict(rec)
        if len(cibles) > 1:
            notes.append(f"« {nom_fiche} » appliquée à {cibles} (mêmes quantités ; "
                         f"affiner les portions PM si besoin).")
        if sauce:
            recettes["SAUCE MIEL AGRUMES (base)"] = sauce
            notes.append("Sauce miel (salade exotique) isolée en sous-recette "
                         "« SAUCE MIEL AGRUMES (base) » — préparation en bulk (72H), "
                         "dose par portion à définir puis à ajouter à SLD GAMBAS AGRUMES.")

    # sous-préparations
    for nom_fiche, canon in SOUS_RECETTES.items():
        feuille = _trouver_feuille(wb, nom_fiche)
        if feuille is None:
            anomalies.append(f"Sous-recette introuvable : « {nom_fiche} ».")
            continue
        rec, _, anos = lire_fiche(wb[feuille])
        anomalies += anos
        if rec:
            recettes[canon] = rec
            notes.append(f"Sous-recette « {canon} » importée (utilisable comme ingrédient).")

    return recettes, anomalies, notes


def appliquer(recettes, chemin_json=RECETTES_JSON, chemin_prov=PROVENANCE_JSON):
    """Fusionne dans recettes_exactes.json (+ provenance) avec sauvegarde horodatée."""
    try:
        with open(chemin_json, encoding="utf-8") as f:
            existantes = json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        existantes = {}
    sauvegarde = None
    if os.path.exists(chemin_json):
        sauvegarde = f"{chemin_json}.bak-{datetime.now():%Y%m%d-%H%M%S}"
        shutil.copyfile(chemin_json, sauvegarde)

    existantes.update(recettes)
    with open(chemin_json, "w", encoding="utf-8") as f:
        json.dump(existantes, f, ensure_ascii=False, indent=2)

    try:
        with open(chemin_prov, encoding="utf-8") as f:
            prov = json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        prov = {}
    for prod in recettes:
        prov[prod] = {"source": "fiches chef Maroc cuisine (xlsx)",
                      "fiche": os.path.basename(XLSX),
                      "score": 1.0,
                      "date": f"{datetime.now():%Y-%m-%d}"}
    with open(chemin_prov, "w", encoding="utf-8") as f:
        json.dump(prov, f, ensure_ascii=False, indent=2)
    return sauvegarde


def main():
    ap = argparse.ArgumentParser(description="Import des fiches cuisine PAUL Maroc.")
    ap.add_argument("--dry-run", action="store_true",
                    help="Affiche ce qui serait importé sans rien écrire.")
    ap.add_argument("xlsx", nargs="?", default=XLSX)
    args = ap.parse_args()

    recettes, anomalies, notes = construire(args.xlsx)

    print("=" * 68)
    print(f"  IMPORT RECETTES CUISINE — {os.path.basename(args.xlsx)}"
          f"{'   [DRY-RUN]' if args.dry_run else ''}")
    print("=" * 68)
    for prod, ings in recettes.items():
        print(f"\n> {prod}  ({len(ings)} ingrédient(s))")
        for ing, q in ings.items():
            print(f"    - {ing:<48} {q}")
    if notes:
        print("\nNotes :")
        for n in notes:
            print(f"  - {n}")
    if anomalies:
        print("\nAnomalies / exclusions :")
        for a in anomalies:
            print(f"  ! {a}")

    if args.dry_run:
        print("\n[DRY-RUN] Rien n'a été écrit. Relance sans --dry-run pour appliquer.")
        return
    if not recettes:
        print("\nRien à importer.")
        return
    sauvegarde = appliquer(recettes)
    print(f"\n[OK] {len(recettes)} entrée(s) écrite(s) dans {os.path.relpath(RECETTES_JSON, _RACINE)}")
    if sauvegarde:
        print(f"     Sauvegarde : {os.path.basename(sauvegarde)}")
    print("     -> Relancer le calcul : python main.py")


if __name__ == "__main__":
    main()
