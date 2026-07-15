# -*- coding: utf-8 -*-
"""
Génère le CLASSEUR EXCEL des recettes à définir avec les chefs.

But : lister tous les produits FABRIQUÉS et ACTIFS (sandwichs, burgers, pains,
viennoiseries, pâtisseries, boissons préparées…) pour que les chefs saisissent
les ingrédients EXACTS de chacun — la base des prévisions matières premières.
Sont exclus (onglet « Exclus », rien n'est supprimé en silence) :
  - les produits COMPOSÉS (famille MENU / Prestation, ou nom contenant
    MENU / FORMULE / COFFRET) : ce sont des assemblages d'autres produits ;
  - les articles REVENDUS tels quels (Coca, Sidi Ali, Kinder…) : pas de recette ;
  - les produits INACTIFS (< SEUIL_UNITES_90J ventes sur les 90 derniers jours).

Chaque feuille-catégorie liste TOUS les produits (sauf revente et mono-stock),
PRÉ-REMPLIS avec la recette actuelle du système : la recette EXACTE
(data/recettes_exactes.json, origine affichée d'après la provenance — recette
chef, fiche technique, extrapolation, estimation) si elle existe, sinon
l'estimation automatique. Corriger est plus rapide que partir de zéro.

Sortie : docs/recettes_produits_a_completer.xlsx
Retour  : outils/importer_recettes_chefs.py (relit le fichier rempli et met à
          jour data/recettes_exactes.json → prévisions matières).

Lancement : python outils/generer_tableau_recettes.py
"""

# --- Script utilitaire : exécutable depuis n'importe où (se cale sur la racine) ---
import os as _os, sys as _sys
_RACINE = _os.path.dirname(_os.path.dirname(_os.path.abspath(__file__)))
_sys.path.insert(0, _RACINE)
_os.chdir(_RACINE)
# ---------------------------------------------------------------------------------

import os
import re

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from paul_forecast import bom, config, couverture, forecast_journalier as fj
from outils import estim_recettes

SORTIE = os.path.join(_RACINE, "docs", "recettes_produits_a_completer.xlsx")

# Provenance des recettes exactes -> libellé « Origine de la ligne » du classeur.
try:
    with open(os.path.join(_RACINE, "data", "recettes_exactes_provenance.json"),
              encoding="utf-8") as _f:
        _PROVENANCE = __import__("json").load(_f)
except (FileNotFoundError, ValueError):
    _PROVENANCE = {}

LIBELLES_SOURCE = {
    "recette chef": "recette chef (validée)",
    "extrapolation recette chef": "extrapolation recette chef",
    "fiche reelle": "fiche technique",
    "tableau chefs (xlsx)": "validée par le chef (tableau)",
    "estimation logique": "estimation (logique produit)",
}

# Notes pré-remplies dans « Commentaire chef » : points précis à trancher.
COMMENTAIRES_PREREMPLIS = {
    ("TARTE PASSION FRAMB", "Fond de tarte"):
        "Chef : 32 g annoncés ; 3,5 kg de pâte / 84 ind = 41,7 g retenus - à vérifier",
    ("TARTE PASSION FRAMB 6P", "Fond de tarte"):
        "Idem individuel : 32 g/part annoncés vs 41,7 g/part retenus - à vérifier",
}

JOURS_ACTIVITE   = 90   # fenêtre d'activité (jours, depuis la dernière vente connue)
SEUIL_UNITES_90J = 5    # en dessous = produit inactif (exclu du tableau)
LIGNES_VIDES_AVEC_ESTIMATION = 3   # lignes vides ajoutées sous une estimation
LIGNES_VIDES_SANS_ESTIMATION = 6   # lignes vides si aucune estimation

FAMILLES_COMPOSEES = {"MENU", "Prestation compl"}
# Composés = menus/formules/coffrets + formules petit-déjeuner (préfixe seulement :
# « MINI CROISS PDJ ENFANT » ou « THE A LA MENTHE PDJ » sont de vrais produits).
MOTIF_COMPOSE = re.compile(r"\bMENU\b|\bFORMULE\b|\bCOFFRET\b"
                           r"|^(?:PDJ|PT DEJ|PTIT DEJ|PETIT DEJ|BRUNCH)\b")
# Marques revendues telles quelles (volontairement SANS « VAE » ni « FRESH » :
# des produits fabriqués les portent, ex. FROM BL NATURE VAE).
MOTS_REVENTE = ("COCA", "SPRITE", "FANTA", "PEPSI", "ORANGINA", "SCHWEPPES",
                "SHWEPS", "7UP", "RED BULL", "REDBULL", "SIDI ALI", "OULMES",
                "EVIAN", "VITTEL", "HAWAI", "ICE TEA", "FRESH UP", "POMS",
                "TWIX", "SNICKERS", "KINDER", "BOUNTY", "KIT KAT", "KITKAT",
                "NESPRESSO", "TASSIMO", "BONNE MAMAN", "EAU MINERALE",
                "EAU MINÉRALE", "EAU PLATE")

# Produits MONO-INGRÉDIENT / déjà en stock : pas de recette à définir.
# Suppléments, accompagnements simples (riz, frites…), boules de glace,
# emballages. (Volontairement APRÈS revente, mais AVANT le reste.)
MOTIF_MONO = re.compile(
    r"\bSUPP|^SUP\b|SUPPL[EÉI]MENT"          # suppléments
    r"|^ACCOMPAGNEMENT"                        # accompagnements (riz, frites, légumes…)
    r"|\bBOULE|POT GLACE"                       # glaces à la boule
    r"|SAC TISSU"                               # emballage (sac tissu)
    r"|^POMMES FRITES$|^FRITES\b|^LEGUMES GRILLES$|^LÉGUMES GRILLÉS$",
    re.I)
MONO_EXACTS = {"FROMAGE", "HUILE D'OLIVE CONFIT"}


def _nature_mono(nom):
    """Libellé court expliquant pourquoi le produit est mono-ingrédient / stock."""
    n = str(nom).upper()
    if re.search(r"\bSUPP|^SUP\b|SUPPL[EÉI]MENT", n):
        return "supplément (article unique déjà en stock)"
    if "BOULE" in n or "POT GLACE" in n:
        return "glace (portion à la boule, déjà en stock)"
    if n.startswith("ACCOMPAGNEMENT") or "FRITES" in n or "LEGUMES GRILLES" in n:
        return "accompagnement (préparé / déjà en stock)"
    if "SAC TISSU" in n:
        return "emballage (non alimentaire)"
    return "article unique (déjà en stock)"

UNITES_VALIDES = ("g", "ml", "unité", "kg", "L", "cl", "pièce")

# ── Palette / styles (sobre, sans emoji) ──────────────────────────────────────
POLICE      = "Arial"
C_ENTETE    = "1C1714"   # brun très foncé
C_CREME     = "F3ECDD"
C_BLOC      = "FBF9F3"   # fond alterné des blocs produits
C_GRIS      = "8A7D6B"
C_OR        = "B8904A"

F_TITRE  = Font(name=POLICE, size=14, bold=True, color=C_ENTETE)
F_TEXTE  = Font(name=POLICE, size=10)
F_GRIS   = Font(name=POLICE, size=10, color=C_GRIS)
F_ENTETE = Font(name=POLICE, size=10, bold=True, color="FFFFFF")
REMPL_ENTETE = PatternFill("solid", start_color=C_ENTETE)
REMPL_BLOC   = PatternFill("solid", start_color=C_BLOC)
BORD_BAS = Border(bottom=Side(style="thin", color="E6DCC7"))


def stats_produits(df_ventes=None):
    """Par produit : famille, ventes des 90 derniers jours, ventes/jour, dernière vente."""
    df = df_ventes if df_ventes is not None else fj.charger_ventes()
    if df is None or df.empty:
        raise SystemExit("Ventes journalières introuvables (donnees_ventes/ventes_journalieres.csv).")
    fin = df["Date"].max()
    rec = df[df["Date"] > fin - pd.Timedelta(days=JOURS_ACTIVITE)]
    g = (df.sort_values("Date").groupby("Produit")
           .agg(Famille=("Famille", "last"), DerniereVente=("Date", "max")))
    g["Total90"] = rec.groupby("Produit")["Quantite"].sum()
    g["Total90"] = g["Total90"].fillna(0.0)
    g["VentesJour"] = (g["Total90"] / JOURS_ACTIVITE).round(1)
    return g.reset_index()


# Score de provenance minimal pour compter une recette comme réellement validée.
# En dessous (ou absente de recettes_exactes_provenance.json, ou provenance sans
# 'score' — ex. "recette proposée (à valider chef)") : toujours "à définir",
# même si le produit a déjà une ligne dans recettes_exactes.json. Sans ce filtre,
# les ~150 recettes qui ne sont que des ESTIMATIONS système (score 0.5) ou des
# propositions jamais validées (ex. FLUTE 250GR) disparaissaient à tort de la
# liste "à valider" alors qu'elles pèsent le plus sur les achats.
SCORE_MIN_VALIDE = 0.85


def _recette_validee(prod):
    """True si la recette de `prod` dans recettes_exactes.json est réellement
    validée par un chef (score de provenance suffisant), pas juste estimée.

    Pas d'entrée de provenance du tout (recettes historiques antérieures au
    suivi de provenance) : on fait confiance, comme avant. Entrée présente
    mais sans 'score' (ex. "recette proposée (à valider chef)") ou score
    insuffisant : recette à définir malgré sa présence dans le BOM.
    """
    entree = _PROVENANCE.get(prod)
    if entree is None:
        return True
    if not isinstance(entree, dict) or "score" not in entree:
        return False
    try:
        return float(entree["score"]) >= SCORE_MIN_VALIDE
    except (TypeError, ValueError):
        return False


def classifier(stats):
    """Ajoute Statut (a_definir | exacte | exclu_compose | exclu_revente | exclu_inactif),
    Source (recette actuelle) et PoidsKgJour (impact matières estimé)."""
    lignes = []
    for _, r in stats.iterrows():
        prod, fam = str(r["Produit"]), str(r["Famille"])
        nom = prod.upper()
        if fam in FAMILLES_COMPOSEES or MOTIF_COMPOSE.search(nom):
            statut, source = "exclu_compose", ""
        elif any(m in nom for m in MOTS_REVENTE):
            statut, source = "exclu_revente", ""
        elif MOTIF_MONO.search(prod) or nom in MONO_EXACTS:
            statut, source = "mono_stock", ""
        elif r["Total90"] < SEUIL_UNITES_90J:
            statut, source = "exclu_inactif", ""
        else:
            source = couverture.source_recette(prod, fam)
            statut = "exacte" if (source == "exacte" and _recette_validee(prod)) else "a_definir"
        poids_g = couverture._poids_matiere_unitaire(prod, fam) if statut in ("a_definir", "exacte") else 0.0
        lignes.append({**r.to_dict(), "Statut": statut, "Source": source,
                       "PoidsKgJour": round(r["VentesJour"] * poids_g / 1000.0, 2)})
    df = pd.DataFrame(lignes)
    return df.sort_values(["Famille", "VentesJour"], ascending=[True, False]).reset_index(drop=True)


def _decompose_ingredient(nom_ing):
    """'Farine de blé T65 (g)' -> ('Farine de blé T65', 'g') ; sans suffixe -> unité 'g'."""
    m = re.match(r"^(.*?)\s*\((g|ml|unité|unite|l|kg|cl|pièce|piece)\)\s*$", str(nom_ing), re.I)
    if not m:
        return str(nom_ing).strip(), "g"
    unite = m.group(2).lower().replace("unite", "unité").replace("piece", "pièce")
    return m.group(1).strip(), ("L" if unite == "l" else unite)


def _recette_estimee(prod, fam):
    """Recette pré-remplie : [(ingrédient, quantité, unité, origine)].

    Priorité : recette EXACTE (recettes_exactes.json, origine = provenance)
    > recette-type enrichie (estim_recettes) > détection par nom (bom)
    > recette générique de famille. Hors recette chef, ce sont des
    ESTIMATIONS à corriger.
    """
    r = config.BOM.get(prod)
    if r:
        src = str(_PROVENANCE.get(prod, {}).get("source", ""))
        origine = LIBELLES_SOURCE.get(src, "recette actuelle (exacte)")
    else:
        r = estim_recettes.estimer(prod, fam)
        origine = "estimation (recette type)"
        if not r:
            r = bom.detecter_bom_produit(prod)
            origine = "estimation (nom du produit)"
        if not r:
            r = bom.recette_generique_famille(prod, fam)
            origine = "estimation (générique famille)"
    out = []
    for ing, q in (r or {}).items():
        try:
            q = float(q)
        except (TypeError, ValueError):
            continue
        nom, unite = _decompose_ingredient(ing)
        out.append((nom, q, unite, origine))
    return out


# ── Construction du classeur ──────────────────────────────────────────────────
def _entetes(ws, titres, largeurs):
    for i, (t, l) in enumerate(zip(titres, largeurs), start=1):
        c = ws.cell(row=1, column=i, value=t)
        c.font, c.fill = F_ENTETE, REMPL_ENTETE
        c.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = l
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 28


def _onglet_mode_emploi(wb, n_a_definir, n_exactes, n_mono, n_exclus, date_donnees):
    ws = wb.active
    ws.title = "Mode d'emploi"
    ws.sheet_properties.tabColor = C_OR
    ws.column_dimensions["A"].width = 108
    textes = [
        ("RECETTES PRODUITS - A COMPLETER PAR LES CHEFS", F_TITRE),
        (f"Généré le {pd.Timestamp.today():%d/%m/%Y} - ventes connues jusqu'au {date_donnees:%d/%m/%Y}. "
         f"{n_a_definir} produits à définir, {n_exactes} déjà définis, {n_mono} mono-ingrédient (sans recette), "
         f"{n_exclus} exclus (menus, revente, inactifs).", F_GRIS),
        ("", F_TEXTE),
        ("POURQUOI : ce classeur alimente le calcul des besoins en matières premières (bon de commande). "
         "Plus les recettes sont exactes, plus les quantités de farine, beurre, garnitures... commandées sont justes.", F_TEXTE),
        ("", F_TEXTE),
        ("COMMENT REMPLIR (UNE FEUILLE PAR CATÉGORIE : BOULANGERIE, CUISINE, PATISSERIE...) :",
         Font(name=POLICE, size=11, bold=True)),
        ("   Chaque catégorie a sa propre feuille (onglets colorés en bas). Chaque chef peut prendre la sienne.", F_TEXTE),
        ("1. Les quantités s'entendent POUR 1 UNITE VENDUE (1 sandwich, 1 baguette, 1 part).", F_TEXTE),
        ("2. Les lignes déjà remplies sont des ESTIMATIONS (composition classique du produit, à VERIFIER) : "
         "corriger les quantités fausses, mettre 0 (ou effacer la quantité) pour supprimer un ingrédient faux.", F_TEXTE),
        ("3. Ajouter les ingrédients manquants dans les lignes vides du produit (colonne Ingrédient + Quantité + Unité).", F_TEXTE),
        ("4. Unités acceptées : g, ml, unité, kg, L, cl, pièce. Les semi-finis sont permis "
         "(ex. Pâte à croissant, Crème pâtissière, Pain nordique) : le système les éclate ensuite en matières de base.", F_TEXTE),
        ("5. L'eau peut être notée (elle est ignorée dans les commandes automatiquement).", F_TEXTE),
        ("6. Quand la recette d'un produit est juste et complète : écrire OUI dans la colonne "
         "« Recette validée (OUI) » sur la LIGNE DU TITRE du produit (la première, celle qui porte le nom). "
         "SEULES LES RECETTES MARQUEES OUI SERONT IMPORTEES.", F_TEXTE),
        ("", F_TEXTE),
        ("EXEMPLE (pour 1 flûte de 250 g) :", Font(name=POLICE, size=11, bold=True)),
        ("   Farine de blé T65 : 180 | g       Eau : 110 | ml       Levure boulangère : 3 | g       Sel : 4,5 | g", F_TEXTE),
        ("", F_TEXTE),
        ("ONGLETS :", Font(name=POLICE, size=11, bold=True)),
        ("- Une feuille par CATÉGORIE (BOULANGERIE, VIENNOISERIE, PATISSERIE, CUISINE, BEVERAGE...) : "
         "TOUS les produits y sont listés (actifs, inactifs, composés, avec recettes exactes ou à définir). "
         "Complétez/corrigez les recettes et marquez OUI sur la ligne-titre de chaque produit fini.", F_TEXTE),
        ("- Mono-ingrédient (stock) : produits sans recette (supplément, boule de glace, riz/frites, emballage) - "
         "RIEN A REMPLIR, ils sont déjà en stock tels quels.", F_TEXTE),
        ("- Revendus (sans recette) : articles achetés tout fait (Coca, Sprite, Kinder...) - "
         "RIEN A REMPLIR.", F_TEXTE),
        ("", F_TEXTE),
        ("NE PAS trier ni supprimer de colonnes. Une fois rempli, redonner le fichier : "
         "l'import se fait avec  python outils/importer_recettes_chefs.py", F_GRIS),
    ]
    for i, (txt, police) in enumerate(textes, start=1):
        c = ws.cell(row=i, column=1, value=txt)
        c.font = police
        c.alignment = Alignment(wrap_text=True, vertical="top")
    return ws


# Ordre d'affichage des feuilles-catégories (les autres familles suivent).
ORDRE_FAMILLES = ["BOULANGERIE", "VIENNOISERIE", "PATISSERIE", "CUISINE",
                  "BEVERAGE", "Autres", "CONFISS/CHOCOLAT"]
# Feuilles NON-recettes (pour l'import : tout le reste est une feuille-catégorie).
FEUILLES_FIXES = ("Mode d'emploi", "Mono-ingrédient (stock)", "Revendus (sans recette)")


def _sanit_sheet(nom):
    """Nom de feuille Excel valide (pas de []:*?/\\, max 31 caractères)."""
    return re.sub(r"[\[\]:*?/\\]", "-", str(nom)).strip()[:31] or "Feuille"


def _onglet_famille(wb, famille, produits):
    """Une feuille de saisie pour UNE catégorie : produits + recettes estimées +
    colonne « Recette validée (OUI) » sur la première ligne de chaque produit."""
    ws = wb.create_sheet(_sanit_sheet(famille))
    ws.sheet_properties.tabColor = C_OR
    _entetes(ws, ["Produit", "Ventes/jour", "Recette validée (OUI)",
                  "Ingrédient", "Quantité (pour 1 unité vendue)", "Unité",
                  "Origine de la ligne", "Commentaire chef"],
             [34, 11, 14, 34, 16, 9, 24, 28])
    dv_u = DataValidation(type="list", formula1='"' + ",".join(UNITES_VALIDES) + '"',
                          allow_blank=True, showErrorMessage=True, errorTitle="Unité invalide",
                          error="Choisir : " + ", ".join(UNITES_VALIDES))
    dv_oui = DataValidation(type="list", formula1='"OUI,NON"', allow_blank=True)
    ws.add_data_validation(dv_u)
    ws.add_data_validation(dv_oui)

    ligne = 2
    for i_prod, (_, p) in enumerate(produits.iterrows()):
        prod, fam = str(p["Produit"]), str(p["Famille"])
        est = _recette_estimee(prod, fam)
        n_vides = LIGNES_VIDES_AVEC_ESTIMATION if est else LIGNES_VIDES_SANS_ESTIMATION
        lignes = est + [("", None, "", "à compléter")] * n_vides
        fond = REMPL_BLOC if i_prod % 2 == 0 else None
        debut = ligne
        for j, (ing, q, unite, origine) in enumerate(lignes):
            prem = (j == 0)
            commentaire = COMMENTAIRES_PREREMPLIS.get((prod, ing), "")
            vals = [prod if prem else "", p["VentesJour"] if prem else None,
                    None, ing, q, unite, origine, commentaire]
            for col, v in enumerate(vals, start=1):
                c = ws.cell(row=ligne, column=col, value=v)
                c.font = F_TEXTE if (prem and col <= 3) or col in (4, 5, 6) else F_GRIS
                if fond:
                    c.fill = fond
                if col == 2:
                    c.number_format = "0.0"
                if col == 5:
                    c.number_format = "0.###"
            ws.cell(row=ligne, column=1).alignment = Alignment(vertical="top")
            ligne += 1
        dv_oui.add(f"C{debut}")                       # OUI sur la ligne-titre du produit
        for col in range(1, 9):
            ws.cell(row=ligne - 1, column=col).border = BORD_BAS
    dv_u.add(f"F2:F{max(ligne - 1, 2)}")
    return ws


def _onglet_simple(wb, titre, colonnes, largeurs, lignes):
    ws = wb.create_sheet(titre)
    _entetes(ws, colonnes, largeurs)
    for i, vals in enumerate(lignes, start=2):
        for col, v in enumerate(vals, start=1):
            c = ws.cell(row=i, column=col, value=v)
            c.font = F_TEXTE
    return ws


def construire_classeur(cls, date_donnees):
    wb = Workbook()
    # Inclure TOUS les produits sauf revente et mono-stock
    n_def = int((cls["Statut"] == "a_definir").sum())
    n_ex = int((cls["Statut"] == "exacte").sum())
    mono = cls[cls["Statut"] == "mono_stock"]
    revente = cls[cls["Statut"] == "exclu_revente"]
    _onglet_mode_emploi(wb, n_def, n_ex, len(mono), len(revente), date_donnees)

    # Une feuille de saisie PAR CATÉGORIE (ordre défini, familles inconnues à la fin).
    # Inclure TOUS les produits SAUF revente (Coca, Sprite...) et emballage/mono-stock.
    # Cela inclut : a_definir, exacte, exclu_inactif, exclu_compose, et permet aux chefs
    # de corriger/compléter n'importe quel produit, même s'il est inactif ou composé.
    a_completer = cls[~cls["Statut"].isin(["exclu_revente", "mono_stock"])]
    familles = sorted(a_completer["Famille"].unique(),
                      key=lambda f: (ORDRE_FAMILLES.index(f) if f in ORDRE_FAMILLES else 99, f))
    for fam in familles:
        prods = a_completer[a_completer["Famille"] == fam].sort_values("VentesJour", ascending=False)
        _onglet_famille(wb, fam, prods)

    _onglet_simple(wb, "Mono-ingrédient (stock)",
                   ["Produit", "Catégorie", "Ventes/jour", "Nature (rien à remplir)"],
                   [36, 15, 11, 40],
                   [[p["Produit"], p["Famille"], p["VentesJour"], _nature_mono(p["Produit"])]
                    for _, p in mono.sort_values("VentesJour", ascending=False).iterrows()])

    _onglet_simple(wb, "Revendus (sans recette)",
                   ["Produit", "Catégorie", "Ventes/jour"],
                   [36, 15, 11],
                   [[p["Produit"], p["Famille"], p["VentesJour"]]
                    for _, p in revente.sort_values("VentesJour", ascending=False).iterrows()])
    return wb


def generer(chemin=SORTIE, df_ventes=None):
    stats = stats_produits(df_ventes)
    cls = classifier(stats)
    df = df_ventes if df_ventes is not None else fj.charger_ventes()
    wb = construire_classeur(cls, df["Date"].max())
    os.makedirs(os.path.dirname(chemin), exist_ok=True)
    try:
        wb.save(chemin)
    except PermissionError:
        # fichier ouvert dans Excel : on écrit une copie horodatée à côté.
        base, ext = os.path.splitext(chemin)
        chemin = f"{base}_{pd.Timestamp.now():%Y%m%d_%H%M}{ext}"
        wb.save(chemin)
        print(f"(Le fichier d'origine était ouvert dans Excel — copie écrite sous un nouveau nom.)")
    n = cls["Statut"].value_counts()
    total_feuilles = cls[~cls["Statut"].isin(["exclu_revente", "mono_stock"])].shape[0]
    print(f"Classeur écrit : {chemin}")
    print(f"  Feuilles-catégories (à compléter/vérifier) : {total_feuilles} produits")
    print(f"    - À définir      : {n.get('a_definir', 0)}")
    print(f"    - Exacte         : {n.get('exacte', 0)}")
    print(f"    - Inactifs       : {n.get('exclu_inactif', 0)}")
    print(f"    - Composés       : {n.get('exclu_compose', 0)} (menus, formules, coffrets)")
    print(f"  Mono-ingrédient (stock) : {n.get('mono_stock', 0)} (suppléments, glaces, riz/frites, emballages)")
    print(f"  Revendus (sans recette) : {n.get('exclu_revente', 0)} (Coca, Sprite, Kinder...)")
    return chemin, cls


if __name__ == "__main__":
    generer()
