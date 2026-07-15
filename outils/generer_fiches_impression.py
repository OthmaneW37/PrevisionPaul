# -*- coding: utf-8 -*-
"""
Génère des FICHES PAPIER condensées (imprimables) listant les produits les
plus vendus SANS recette exacte, avec de l'espace pour que le chef écrive
à la main les ingrédients / quantités.

Différence avec outils/generer_tableau_recettes.py : ce classeur-là liste
TOUS les produits avec une estimation pré-remplie à corriger (exhaustif,
pensé pour une saisie à l'écran). Celui-ci est volontairement COURT — les
N produits qui pèsent le plus sur les achats par catégorie, sans rien de
pré-rempli, pour être imprimé et rempli au stylo en cuisine.

Sortie : docs/fiches_recettes_a_remplir.xlsx (une feuille par catégorie,
mise en page portrait prête à imprimer).

Lancement : python outils/generer_fiches_impression.py [--top N]
"""

# --- Script utilitaire : exécutable depuis n'importe où (se cale sur la racine) ---
import os as _os, sys as _sys
_RACINE = _os.path.dirname(_os.path.dirname(_os.path.abspath(__file__)))
_sys.path.insert(0, _RACINE)
_os.chdir(_RACINE)
# ---------------------------------------------------------------------------------

import argparse
import os

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

from outils.generer_tableau_recettes import stats_produits, classifier, ORDRE_FAMILLES

SORTIE = os.path.join(_RACINE, "docs", "fiches_recettes_a_remplir.xlsx")

POLICE   = "Arial"
C_ENTETE = "1C1714"
C_OR     = "B8904A"
F_TITRE  = Font(name=POLICE, size=13, bold=True, color=C_ENTETE)
F_SOUS   = Font(name=POLICE, size=9, color="8A7D6B")
F_PROD   = Font(name=POLICE, size=11, bold=True)
F_VENTE  = Font(name=POLICE, size=9, color="8A7D6B")
F_ENTETE = Font(name=POLICE, size=9, bold=True, color="FFFFFF")
REMPL_ENTETE = PatternFill("solid", start_color=C_ENTETE)
LIGNE_ECRITURE = Border(bottom=Side(style="dotted", color="B8904A"))
LIGNE_SEP = Border(top=Side(style="thin", color="1C1714"))

LIGNES_A_ECRIRE = 4   # lignes pointillées libres par produit (ingrédient + quantité)


def _feuille_categorie(wb, famille, produits, top_n):
    ws = wb.create_sheet(famille[:31])
    ws.sheet_properties.tabColor = C_OR
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.6, bottom=0.5)
    ws.print_title_rows = "1:3"

    largeurs = [4, 30, 11, 30, 16]
    for i, l in enumerate(largeurs, start=1):
        ws.column_dimensions[get_column_letter(i)].width = l

    ws.merge_cells("A1:E1")
    c = ws.cell(row=1, column=1, value=f"{famille} — recettes à compléter (produits les plus vendus)")
    c.font = F_TITRE
    ws.merge_cells("A2:E2")
    c = ws.cell(row=2, column=1, value="Quantités POUR 1 UNITÉ VENDUE. Une ligne = un ingrédient + sa quantité "
                                        "(g, ml, unité...). Rayer l'ingrédient s'il ne s'applique pas.")
    c.font = F_SOUS
    c.alignment = Alignment(wrap_text=True)
    ws.row_dimensions[2].height = 24

    entetes = ["#", "Produit", "Ventes/jour", "Ingrédient", "Quantité + unité"]
    for i, t in enumerate(entetes, start=1):
        cell = ws.cell(row=3, column=i, value=t)
        cell.font, cell.fill = F_ENTETE, REMPL_ENTETE
    ws.freeze_panes = "A4"

    sel = produits.sort_values("VentesJour", ascending=False).head(top_n)
    ligne = 4
    for rang, (_, p) in enumerate(sel.iterrows(), start=1):
        debut = ligne
        for j in range(LIGNES_A_ECRIRE):
            if j == 0:
                ws.cell(row=ligne, column=1, value=rang).font = F_PROD
                ws.cell(row=ligne, column=2, value=str(p["Produit"])).font = F_PROD
                ws.cell(row=ligne, column=3, value=round(float(p["VentesJour"]), 1)).font = F_VENTE
            for col in (4, 5):
                cell = ws.cell(row=ligne, column=col)
                cell.border = LIGNE_ECRITURE
            ws.row_dimensions[ligne].height = 16
            ligne += 1
        for col in range(1, 6):
            ws.cell(row=ligne - 1, column=col).border = Border(
                bottom=Side(style="thin", color="1C1714"))
    return ws


def generer(chemin=SORTIE, top_n=20, df_ventes=None):
    stats = stats_produits(df_ventes)
    cls = classifier(stats)
    a_definir = cls[cls["Statut"] == "a_definir"]

    wb = Workbook()
    wb.remove(wb.active)
    familles = sorted(a_definir["Famille"].unique(),
                      key=lambda f: (ORDRE_FAMILLES.index(f) if f in ORDRE_FAMILLES else 99, f))
    total = 0
    for fam in familles:
        prods = a_definir[a_definir["Famille"] == fam]
        if prods.empty:
            continue
        _feuille_categorie(wb, fam, prods, top_n)
        total += min(top_n, len(prods))

    os.makedirs(os.path.dirname(chemin), exist_ok=True)
    try:
        wb.save(chemin)
    except PermissionError:
        base, ext = os.path.splitext(chemin)
        import pandas as pd
        chemin = f"{base}_{pd.Timestamp.now():%Y%m%d_%H%M}{ext}"
        wb.save(chemin)
    print(f"Fiches écrites : {chemin}")
    print(f"  {len(familles)} catégories, {total} produits au total (top {top_n}/catégorie)")
    return chemin


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--top", type=int, default=20, help="nb de produits par catégorie (défaut 20)")
    args = ap.parse_args()
    generer(top_n=args.top)
