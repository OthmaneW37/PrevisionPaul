# -*- coding: utf-8 -*-
"""
Importe le classeur « recettes_produits_a_completer.xlsx » REMPLI par les chefs
dans data/recettes_exactes.json (+ provenance), pour fiabiliser les prévisions
matières premières.

Règles :
  - le classeur a UNE FEUILLE PAR CATÉGORIE (BOULANGERIE, CUISINE...) ; on lit
    toutes ces feuilles (tout sauf Mode d'emploi / Mono-ingrédient / Revendus,
    et les anciennes feuilles Déjà définies / Exclus) ;
  - seuls les produits marqués OUI (colonne « Recette validée (OUI) », sur la
    ligne-titre du produit) sont importés ;
  - leurs lignes avec Ingrédient + Quantité > 0 deviennent la recette EXACTE
    (elle remplace l'ancienne entrée du produit) ;
  - unités converties : kg -> g (x1000), L -> ml (x1000), cl -> ml (x10),
    pièce -> unité ; l'ingrédient est stocké « Nom (g|ml|unité) » ;
  - une sauvegarde horodatée de recettes_exactes.json est écrite avant toute
    modification ; un rapport liste produits importés et anomalies.

Après import : relancer le calcul (python main.py + forecast journalier) pour
répercuter sur le bon de commande. Penser à retirer le calibrage provisoire
CALIBRAGE_MATIERES (paul_forecast/config.py) quand les recettes clés sont vraies.

Lancement : python outils/importer_recettes_chefs.py [chemin_du_fichier.xlsx]
"""

# --- Script utilitaire : exécutable depuis n'importe où (se cale sur la racine) ---
import os as _os, sys as _sys
_RACINE = _os.path.dirname(_os.path.dirname(_os.path.abspath(__file__)))
_sys.path.insert(0, _RACINE)
_os.chdir(_RACINE)
# ---------------------------------------------------------------------------------

import json
import os
import shutil
import sys
from datetime import datetime

from openpyxl import load_workbook

DEFAUT_XLSX = os.path.join(_RACINE, "docs", "recettes_produits_a_completer.xlsx")
RECETTES_JSON = os.path.join(_RACINE, "data", "recettes_exactes.json")
PROVENANCE_JSON = os.path.join(_RACINE, "data", "recettes_exactes_provenance.json")

# (facteur vers l'unité de base, unité de base)
CONVERSIONS = {"g": (1, "g"), "kg": (1000, "g"),
               "ml": (1, "ml"), "l": (1000, "ml"), "cl": (10, "ml"),
               "unité": (1, "unité"), "unite": (1, "unité"),
               "pièce": (1, "unité"), "piece": (1, "unité")}

# Feuilles qui ne sont PAS des feuilles-catégories de saisie.
# (« Déjà définies » / « Exclus » : anciens classeurs, gardés par compatibilité.)
FEUILLES_NON_RECETTES = {"Mode d'emploi", "Mono-ingrédient (stock)",
                         "Revendus (sans recette)", "Déjà définies", "Exclus"}


def _nombre(v):
    """Cellule quantité -> float (accepte '12,5'). None si vide/illisible."""
    if v is None or (isinstance(v, str) and not v.strip()):
        return None
    try:
        return float(str(v).replace(",", ".").strip())
    except ValueError:
        return None


def lire_classeur(chemin):
    """Lit le classeur rempli -> (recettes {produit: {ingrédient: qté}}, anomalies).

    Parcourt chaque feuille-catégorie. Colonnes : A Produit, B Ventes/jour,
    C Validée (OUI), D Ingrédient, E Quantité, F Unité.
    """
    wb = load_workbook(chemin, data_only=True)
    recettes, anomalies, valides = {}, [], set()
    for nom_ws in wb.sheetnames:
        if nom_ws in FEUILLES_NON_RECETTES:
            continue
        ws = wb[nom_ws]
        prod, actif = None, False
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row:
                continue
            titre = str(row[0]).strip() if row[0] else ""
            if titre:                                  # ligne-titre d'un produit
                prod = titre
                marque = str(row[2] or "").strip().upper() if len(row) > 2 else ""
                actif = marque in ("OUI", "O", "X")
                if actif:
                    valides.add(prod)
            if not prod or not actif:
                continue
            ing = str(row[3] or "").strip() if len(row) > 3 else ""
            qte = _nombre(row[4] if len(row) > 4 else None)
            unite = str(row[5] or "").strip().lower() if len(row) > 5 else ""
            if not ing or qte is None:
                continue                               # ligne vide / non remplie
            if qte <= 0:
                continue                               # 0 = ingrédient supprimé
            if unite not in CONVERSIONS:
                anomalies.append(f"{nom_ws} L{i} ({prod}) : unité « {row[5]} » inconnue - ligne ignorée")
                continue
            facteur, base = CONVERSIONS[unite]
            recettes.setdefault(prod, {})[f"{ing} ({base})"] = round(qte * facteur, 2)
    for prod in sorted(valides - set(recettes)):
        anomalies.append(f"« {prod} » marqué OUI mais aucune ligne d'ingrédient exploitable - non importé")
    return recettes, anomalies


def importer(chemin_xlsx=DEFAUT_XLSX, chemin_json=RECETTES_JSON,
             chemin_prov=PROVENANCE_JSON):
    """Importe le classeur -> met à jour recettes exactes + provenance. Rapport dict."""
    recettes, anomalies = lire_classeur(chemin_xlsx)
    if not recettes:
        return {"importes": 0, "produits": [], "anomalies": anomalies, "sauvegarde": None}

    try:
        with open(chemin_json, encoding="utf-8") as f:
            existantes = json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        existantes = {}
    sauvegarde = f"{chemin_json}.bak-{datetime.now():%Y%m%d-%H%M%S}"
    if os.path.exists(chemin_json):
        shutil.copyfile(chemin_json, sauvegarde)
    else:
        sauvegarde = None

    existantes.update(recettes)          # la recette du chef remplace l'ancienne
    with open(chemin_json, "w", encoding="utf-8") as f:
        json.dump(existantes, f, ensure_ascii=False, indent=2)

    try:
        with open(chemin_prov, encoding="utf-8") as f:
            prov = json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        prov = {}
    for prod in recettes:
        prov[prod] = {"source": "tableau chefs (xlsx)",
                      "fiche": os.path.basename(chemin_xlsx),
                      "score": 1.0,
                      "date": f"{datetime.now():%Y-%m-%d}"}
    with open(chemin_prov, "w", encoding="utf-8") as f:
        json.dump(prov, f, ensure_ascii=False, indent=2)

    return {"importes": len(recettes), "produits": sorted(recettes),
            "anomalies": anomalies, "sauvegarde": sauvegarde}


if __name__ == "__main__":
    chemin = sys.argv[1] if len(sys.argv) > 1 else DEFAUT_XLSX
    r = importer(chemin)
    print(f"{r['importes']} recette(s) importée(s) depuis {chemin}")
    for p in r["produits"]:
        print(f"  - {p}")
    if r["sauvegarde"]:
        print(f"Sauvegarde de l'ancien fichier : {r['sauvegarde']}")
    if r["anomalies"]:
        print("Anomalies :")
        for a in r["anomalies"]:
            print(f"  ! {a}")
    if r["importes"]:
        print("Relancer le calcul (python main.py) pour mettre à jour le bon de commande. "
              "Si les recettes clés (farine) sont désormais exactes, retirer CALIBRAGE_MATIERES "
              "dans paul_forecast/config.py.")
