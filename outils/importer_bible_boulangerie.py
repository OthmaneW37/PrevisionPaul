# -*- coding: utf-8 -*-
"""
Import des fiches techniques « Bible Boulangerie » (docs/Boulangerie PAUL.xlsx)
dans data/recettes_exactes.json (+ provenance).

Le classeur contient les tableaux de fabrication des pâtes (kg d'ingrédients
par kg de pâte — tableaux linéaires) et un tableau façonnage/cuisson donnant
le poids de pâte crue par unité pour le pain complet. Ce script :

  1. normalise chaque pâte en ratios (g d'ingrédient / g de pâte) ;
  2. décompose les semi-finis internes (pâte fermentée, levains) en
     farine/eau/sel/levure — hypothèses standard documentées ;
  3. croise pâte × poids cru par produit VENDU (poids façonnage chef quand la
     fiche le donne, sinon grammage du nom + perte de cuisson = HYPOTHÈSE) ;
  4. met à jour recettes_exactes.json + recettes_exactes_provenance.json
     (sauvegarde horodatée des deux fichiers avant écriture).

Relancer après toute modification du classeur :
    python outils/importer_bible_boulangerie.py
puis régénérer les besoins matières : python main.py
"""
import json
import os
import shutil
import sys
from datetime import datetime

from openpyxl import load_workbook

_RACINE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CLASSEUR = os.path.join(_RACINE, "docs", "Boulangerie_PAUL.xlsx")
RECETTES_JSON = os.path.join(_RACINE, "data", "recettes_exactes.json")
PROVENANCE_JSON = os.path.join(_RACINE, "data", "recettes_exactes_provenance.json")

# ── Noms de matières : alignés sur data/alias_matieres.json pour que le MRP
#    agrège ces recettes avec les existantes (et que les prix s'appliquent).
RENOMMAGE = {
    "Semoule fine": "Semoule fine",
    "Sel": "Sel",
    "Levure fraîche": "Levure fraiche",
    "Levure": "Levure fraiche",
    "Améliorant Ibis Bleu": "Ameliorant Ibis Bleu",
    "Améliorant": "Ameliorant Ibis Bleu",
    "Eau vrac": "Eau",
    "Eau": "Eau",
    "Sucre semoule": "Sucre",
    "Sucre": "Sucre",
    "Farine Soisson": "Farine T55",
    "Farine blanche T55 Soisson": "Farine T55",
    "Farine boulangère Soisson": "Farine T55",
    "Farine orge": "Farine d'orge",
    "Farine pain maison": "Farine pain maison",
    "Sésame doré local": "Graines de sesame",
    "Anis graines": "Graines d'anis",
    "Lait entier UHT": "Lait",
    "Beurre": "Beurre",
    "Beurre pâtissier": "Beurre",
    "Farine complète": "Farine complete",
    "Mélange 5 graines": "Melange graines",
    "Farine 6 céréales 4 graines": "Farine 6 cereales 4 graines",
    "Farine malt-orge torréfié": "Farine malt-orge torrefie",
    "Farine viennoiserie": "Farine T45",
    "Œufs frais": "Œufs",
    "Chocolat goutte": "Pepites de chocolat",
    "Origan sec": "Origan",
}

# ── Semi-finis internes aux pâtes, décomposés en matières de base.
#    HYPOTHÈSES standard boulangerie (à affiner avec le chef si besoin) :
#    - pâte fermentée = pâte blanche de la veille (composition pain blanc,
#      point fixe du tableau « Pate pain blanc ») ;
#    - levains = 50 % farine + 50 % eau (levain liquide / rafraîchi).
DECOMPOSITION_SEMI_FINIS = {
    "Pâte fermentée": {"Farine T55": 0.604, "Eau": 0.381, "Levure fraiche": 0.012, "Sel": 0.012},
    "Levain rafraîchi": {"Farine T55": 0.50, "Eau": 0.50},
    "Levain liquide": {"Farine T55": 0.50, "Eau": 0.50},
    "Pâte pain de mie": None,   # résolu dynamiquement (ratios de la feuille)
}


def _ratios_feuille(ws):
    """Ratios {matière: g par g de pâte} depuis un tableau de fabrication.

    Les tableaux sont linéaires : on prend la ligne au poids de pâte le plus
    élevé (moins d'erreur d'arrondi) et on divise chaque colonne par ce poids.
    Certaines feuilles (brioche/burger) donnent des colonnes en GRAMMES pour
    une farine en KG : détecté via le ratio (> 1 impossible en kg/kg).
    """
    lignes = list(ws.iter_rows(min_row=1, values_only=True))
    # L'en-tête est la première ligne dont la 1re cellule commence par
    # « Poids » ou « Total » (les feuilles ont un titre + parfois une ligne vide).
    i_entete = next(i for i, l in enumerate(lignes)
                    if str(l[0] or "").strip().startswith(("Poids", "Total")))
    entetes = [str(v).strip() for v in lignes[i_entete] if v is not None]
    meilleures = None
    for vals in lignes[i_entete + 1:]:
        if vals[0] is None or not isinstance(vals[0], (int, float)):
            continue        # fin du tableau (procédé / notes)
        if meilleures is None or vals[0] > meilleures[0]:
            meilleures = vals
    poids = float(meilleures[0])
    ratios = {}
    for i, ent in enumerate(entetes[1:], start=1):
        v = meilleures[i]
        if v is None:
            continue
        nom = ent.split("(")[0].strip()      # retire « (kg) », « (g) 2,50% »…
        ratio = float(v) / poids
        if "(g)" in ent:                      # colonnes en grammes
            ratio /= 1000.0
        ratios[nom] = ratios.get(nom, 0.0) + ratio
    return ratios


def _decomposer(ratios, pates):
    """Remplace les semi-finis (pâte fermentée, levains, pâte pain de mie)
    par leurs matières de base, et renomme vers les alias canoniques."""
    out = {}
    for nom, r in ratios.items():
        if nom in DECOMPOSITION_SEMI_FINIS:
            sous = DECOMPOSITION_SEMI_FINIS[nom] or pates["Pâte pain de mie"]
            for m, rr in sous.items():
                out[m] = out.get(m, 0.0) + r * rr
        else:
            m = RENOMMAGE.get(nom, nom)
            out[m] = out.get(m, 0.0) + r
    return out


def charger_pates(classeur=CLASSEUR):
    """Toutes les pâtes du classeur en ratios base-matières {pâte: {matière: ratio}}."""
    wb = load_workbook(classeur, data_only=True)
    feuilles = {
        "Pâte semoule": "Pate semoule",
        "Pâte pain orge": "Pate pain orge",
        "Pâte traditionnelle": "Pate traditionnel",
        "Pâte pain blanc": "Pate pain blanc",
        "Pâte pain de mie": "Pate pain de mie",
        "Pâte pain de mie complet": "Pate pain de mie complet",
        "Pâte moulé complet grainé": "Moule complet graine",
        "Pâte nordique": "Pate nordique",
        "Pâte 6 céréales grainé": "Pate 6 cereales graine",
        "Pâte à brioche": "Pate a brioche",
        "Burger brioche": "Burger brioche 100g",
        "Hamburger origan": "Hamburger origan 110g",
        "Viennoise pépites choco": "Viennoise pepites choco",
        "Pâte à l'ancienne": "Pate a l" + "'" + "ancienne",
        "Pâte complète": "Pate complete",
    }
    brutes = {nom: _ratios_feuille(wb[f]) for nom, f in feuilles.items()}
    # « Pâte pain de mie » d'abord : les dérivés (hamburger, viennoise) en dépendent.
    pates = {}
    pates["Pâte pain de mie"] = _decomposer(brutes.pop("Pâte pain de mie"), {})
    for nom, ratios in brutes.items():
        pates[nom] = _decomposer(ratios, pates)
    # Note : « Pâte complète » est maintenant dans la feuille « Pate complete »
    # (pâte avec levain). Supprimer la dérivation d'hypothèse si elle apparaît.
    return pates


# ── Produits VENDUS → (pâte, poids de pâte crue en g, extras, source du poids).
#    Poids « façonnage chef » = feuille « Faconnage cuisson complet » (63 g
#    pistolet, 84 g pain rond, 310 g flûtes 200-300 g, 495 g formats 400-500 g)
#    appliquée aussi aux autres pâtes (mêmes formats de façonnage).
#    Poids « hypothèse » = grammage du nom / (1 - perte de cuisson) :
#    pain 20 %, brioche/viennoiserie 15 %, moulé pain de mie 10 %.
CHEF = "façonnage chef"
HYP = lambda d: f"hypothèse : {d}"

PRODUITS = {
    # ── Pâte pain blanc ────────────────────────────────────────────────────
    "FLUTE 250GR":            ("Pâte pain blanc", 310, {}, CHEF),
    "FLUTE 200GR":            ("Pâte pain blanc", 310, {}, CHEF),
    "FLUTE NATURE 200GR":     ("Pâte pain blanc", 310, {}, CHEF),
    "FLUTE NATURE 300GR":     ("Pâte pain blanc", 310, {}, CHEF),
    "FLUTE SS SEL 250GR":     ("Pâte pain blanc", 310, {"__sans_sel__": True}, CHEF),
    "FLUTE NATURE 400GR":     ("Pâte pain blanc", 495, {}, CHEF),
    "FLUTE OLIVE 250GR":      ("Pâte pain blanc", 310, {"Olives denoyautees": 35.0}, CHEF),
    "FLUTE SESAME 200GR":     ("Pâte pain blanc", 310, {"Graines de sesame": 12.0}, CHEF),
    "FLUTE PAVOT 200GR":      ("Pâte pain blanc", 310, {"Graines de pavot": 10.0}, CHEF),
    "Chapata (pcf)":          ("Pâte pain blanc", 310, {}, CHEF),
    "PAIN ESPAGNOL 100GR":    ("Pâte pain blanc", 125, {}, HYP("100 g cuit, perte 20 %")),
    "PAIN ROND BLANC 200GR":  ("Pâte pain blanc", 250, {}, HYP("200 g cuit, perte 20 %")),
    "PISTOLET NATURE 50GR":   ("Pâte pain blanc", 63, {}, CHEF),
    "PISTOLET OLIVE 50GR":    ("Pâte pain blanc", 63, {"Olives denoyautees": 8.0}, CHEF),
    "PISTOLET SESAME 50GR":   ("Pâte pain blanc", 63, {"Graines de sesame": 3.0}, CHEF),
    "PISTOLET PAVOT 50GR":    ("Pâte pain blanc", 63, {"Graines de pavot": 2.0}, CHEF),
    "MINI PAIN LONG 50GR":    ("Pâte pain blanc", 63, {}, CHEF),
    "MINI PAIN ROND 50GR":    ("Pâte pain blanc", 63, {}, CHEF),
    # ── Semoule / orge / traditionnel ─────────────────────────────────────
    "PAIN PUR SEMOULE 80GR":  ("Pâte semoule", 100, {}, HYP("80 g cuit, perte 20 %")),
    "PAIN ROND SEMOULE 220GR": ("Pâte semoule", 275, {}, HYP("220 g cuit, perte 20 %")),
    "BAGUETTE MIX SEMOULE":   ("Pâte semoule", 310, {}, HYP("format flûte")),
    "PAIN ORGE 90G":          ("Pâte pain orge", 113, {}, HYP("90 g cuit, perte 20 %")),
    "PAIN TRADITIONNEL 200GR": ("Pâte traditionnelle", 250, {}, HYP("200 g cuit, perte 20 %")),
    # ── Complet (moulé complet grainé) ────────────────────────────────────
    "PAIN COMPLET 80GR":      ("Pâte moulé complet grainé", 84, {}, CHEF),
    "PISTOLET COMPLET 50GR":  ("Pâte moulé complet grainé", 63, {"Flocons d'orge": 2.0}, CHEF),
    "FLUTE COMPLETE 200GR":   ("Pâte moulé complet grainé", 310, {}, CHEF),
    "FLUTE COMPLETE SS SEL 200": ("Pâte moulé complet grainé", 310, {"__sans_sel__": True}, CHEF),
    "PAIN COMPLET 400G":      ("Pâte moulé complet grainé", 495, {}, CHEF),
    "MOULE COMPLET 400GR":    ("Pâte moulé complet grainé", 495, {}, CHEF),
    "PAIN COMPLET SON 400GR": ("Pâte moulé complet grainé", 495, {"Son de ble": 15.0}, CHEF),
    "MOULE COMPLET GRAINE":   ("Pâte moulé complet grainé", 495, {}, CHEF),
    # ── Nordique / 6 céréales ─────────────────────────────────────────────
    "PAIN NORDIQUE 400G":     ("Pâte nordique", 495, {}, CHEF),
    "Baguette nordique":      ("Pâte nordique", 310, {}, HYP("format flûte")),
    "FLUTE GRAINEE 250G":     ("Pâte 6 céréales grainé", 310, {}, CHEF),
    "PAIN GRAINE 400G":       ("Pâte 6 céréales grainé", 495, {}, CHEF),
    # ── Pain de mie / viennoise / hamburger ───────────────────────────────
    "MOULE FERME 750GR":      ("Pâte pain de mie", 833, {}, HYP("750 g cuit, perte 10 %")),
    "PAIN MIE COMPLET 750GR": ("Pâte pain de mie complet", 833, {}, HYP("750 g cuit, perte 10 %")),
    "VIENNOISE NATURE 100GR": ("Pâte pain de mie", 118, {}, HYP("100 g cuit, perte 15 %")),
    "Pain hamburger (pcf)":   ("Hamburger origan", 110, {}, "fiche (110 g cru)"),
    "Mini pain hamburger (pcf)": ("Hamburger origan", 55, {}, HYP("moitié du format 110 g")),
    # ── Brioche ────────────────────────────────────────────────────────────
    "BRIOCHE 300GR":          ("Pâte à brioche", 353, {}, HYP("300 g cuit, perte 15 %")),
    "Brioche tresse 300gr":   ("Pâte à brioche", 353, {}, HYP("300 g cuit, perte 15 %")),
    "SAC MINI PAIN BURGER 4PCS": ("Burger brioche", 240, {}, HYP("4 × 60 g cru")),
    # ── Pâte à l'ancienne (hydratation élevée, levure faible) ──────────────
    "BENOITON FROMAGE":       ("Pâte à l'ancienne", 70, {"Fromage": 10.0}, CHEF),
    "BENOITON OLIVES":        ("Pâte à l'ancienne", 70, {"Olives denoyautees": 10.0}, CHEF),
    "BENOITON OLIVE SÉSAME":  ("Pâte à l'ancienne", 70, {"Olives denoyautees": 5.0, "Graines de sesame": 5.0}, CHEF),
    "Paulette fromage (pcf)": ("Pâte à l'ancienne", 140, {"Fromage": 20.0}, CHEF),
    "Paulette sésame (pcf)":  ("Pâte à l'ancienne", 140, {"Graines de sesame": 10.0}, CHEF),
    "Paulette pavot (pcf)":   ("Pâte à l'ancienne", 140, {"Graines de pavot": 8.0}, CHEF),
    # Pistolets 6 céréales (pâte 6 céréales, 63 g blanc)
    "Pistolet 6 céré 50gr":   ("Pâte 6 céréales grainé", 63, {}, CHEF),
    # Flûtes 6 céréales, ancienne
    "Flute 6Ceriale 200gr":   ("Pâte 6 céréales grainé", 310, {}, CHEF),
}


def construire_recettes(pates):
    """Recettes {produit: {matière: g/unité}} + provenance {produit: dict}."""
    recettes, provenances = {}, {}
    for produit, (pate, poids, extras, source_poids) in PRODUITS.items():
        ratios = pates[pate]
        rec = {m: round(r * poids, 2) for m, r in ratios.items() if r * poids >= 0.05}
        sans_sel = extras.pop("__sans_sel__", False) if isinstance(extras, dict) else False
        if sans_sel:
            rec.pop("Sel", None)
        for m, g in extras.items():
            rec[m] = rec.get(m, 0.0) + g
        recettes[produit] = rec
        provenances[produit] = {
            "source": "fiche reelle",
            "fiche": f"Bible Boulangerie — {pate}",
            "poids_pate_g": poids,
            "poids_source": source_poids,
        }
    return recettes, provenances


def importer(classeur=CLASSEUR, chemin_rec=RECETTES_JSON, chemin_prov=PROVENANCE_JSON):
    """Importe le classeur → met à jour recettes exactes + provenance."""
    pates = charger_pates(classeur)
    recettes, provenances = construire_recettes(pates)

    with open(chemin_rec, encoding="utf-8") as f:
        rec_exist = json.load(f)
    with open(chemin_prov, encoding="utf-8") as f:
        prov_exist = json.load(f)

    horodatage = datetime.now().strftime("%Y%m%d_%H%M%S")
    for chemin in (chemin_rec, chemin_prov):
        shutil.copy2(chemin, f"{chemin}.bak_{horodatage}")

    remplaces = [p for p in recettes if p in rec_exist]
    ajoutes = [p for p in recettes if p not in rec_exist]
    rec_exist.update(recettes)
    prov_exist.update(provenances)

    with open(chemin_rec, "w", encoding="utf-8") as f:
        json.dump(rec_exist, f, ensure_ascii=False, indent=2)
    with open(chemin_prov, "w", encoding="utf-8") as f:
        json.dump(prov_exist, f, ensure_ascii=False, indent=2)

    return {"pates": pates, "remplaces": remplaces, "ajoutes": ajoutes,
            "sauvegarde": horodatage}


if __name__ == "__main__":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    res = importer()
    print(f"Sauvegarde des JSON existants : *.bak_{res['sauvegarde']}")
    print(f"\nPâtes normalisées ({len(res['pates'])}) — g de matière par g de pâte :")
    for nom, ratios in res["pates"].items():
        detail = ", ".join(f"{m} {r:.3f}" for m, r in
                           sorted(ratios.items(), key=lambda x: -x[1]))
        print(f"  [{nom}] {detail}")
    print(f"\nRecettes REMPLACÉES ({len(res['remplaces'])}) : {', '.join(res['remplaces'])}")
    print(f"\nRecettes AJOUTÉES ({len(res['ajoutes'])}) : {', '.join(res['ajoutes'])}")
    print("\nPenser à relancer : python main.py (besoins matières + plan).")
